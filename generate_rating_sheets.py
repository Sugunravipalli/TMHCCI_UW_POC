"""
Generate 50 realistic Rating Data Sheet Excel workbooks for Tokio Marine HCC
commercial property underwriting POC.

Each workbook has 4 sheets: Summary, Rating, Loss Analysis, UW Notes.
Distribution: 10 hero, 15 bound, 10 declined, 8 NTU, 5 referred, 2 renewal-deteriorating.
"""

import os
import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "07_rating_data_sheets")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
VALUE_FONT = Font(name="Calibri", size=10)
NOTE_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT_W = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT_W
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = VALUE_FONT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt

def write_label_value(ws, row, label, value, start_col=1, fmt=None):
    c1 = ws.cell(row=row, column=start_col, value=label)
    c1.font = LABEL_FONT
    c1.border = THIN_BORDER
    c2 = ws.cell(row=row, column=start_col + 1, value=value)
    c2.font = VALUE_FONT
    c2.border = THIN_BORDER
    if fmt:
        c2.number_format = fmt

# ---------------------------------------------------------------------------
# Data pools for random generation
# ---------------------------------------------------------------------------
UNDERWRITERS = ["Simon Thornton", "James Chen", "Rebecca Hartley", "David Kim",
                "Sarah Mitchell", "Mark Stevens"]
BROKERS_FIRMS = {
    "Aon": ["Paul Henderson", "Lisa Chang", "David Morris"],
    "Marsh": ["James Whitfield", "Sarah Price", "Tom Bradley"],
    "Willis": ["Claire Donovan", "Robert Chen", "Emily Watson"],
    "Lockton": ["Emily Richards", "Jason Brooks", "Amanda Foster"],
    "Howden": ["Michael Cross", "Helen Taylor", "Ian Russell"],
    "Gallagher": ["Rachel Foster", "Kevin O'Brien", "Diane Murphy"],
    "McGill": ["Thomas Keane", "Fiona Campbell", "Derek Stone"],
    "Miller Insurance": ["Charlotte Webb", "Ben Archer", "Lucy Ward"],
    "BMS Group": ["Andrew Patel", "Samantha Lee", "Marcus Wright"],
    "Ed Broking": ["Laura Simpson", "Greg Thomson", "Natalie Ford"],
    "Guy Carpenter": ["Neil Armstrong", "Sophie Jenkins", "Richard Hale"],
}

OCCUPANCY_TYPES = [
    "Commercial Office", "Light Manufacturing", "Heavy Manufacturing",
    "Warehouse / Distribution", "Retail", "Healthcare / Hospital",
    "Pharmaceutical", "Data Centre", "Hotel / Hospitality",
    "Food Processing", "Educational", "Automotive",
    "Agricultural", "Mixed-Use Commercial", "Self-Storage",
    "Cold Storage / Refrigeration", "Printing / Publishing",
    "Chemical Processing", "Electronics Assembly", "Real Estate Portfolio",
]

CONSTRUCTION_TYPES = [
    "Fire Resistive (ISO Class 6)", "Non-Combustible (ISO Class 4)",
    "Joisted Masonry (ISO Class 3)", "Masonry Non-Combustible (ISO Class 5)",
    "Frame (ISO Class 1)", "Modified Fire Resistive (ISO Class 5)",
    "Heavy Timber (ISO Class 2)",
]

TERRITORIES = {
    "Northeast US": 1.00, "Southeast US": 1.15, "Midwest US": 0.95,
    "Southwest US": 1.10, "Pacific Northwest US": 0.90, "Florida": 1.35,
    "Gulf Coast US": 1.30, "California": 1.20, "Mountain West US": 0.85,
    "Central London": 1.05, "South East England": 1.00, "Midlands UK": 0.90,
    "North West England": 0.88, "Scotland": 0.85, "Greater Manchester": 0.92,
    "Texas Gulf": 1.28, "Mid-Atlantic US": 1.02, "Caribbean": 1.50,
}


# ---------------------------------------------------------------------------
# HERO accounts -- 10 hand-crafted records
# ---------------------------------------------------------------------------
HERO_ACCOUNTS = [
    # 1 - Pacific Retail
    {
        "filename": "RS_pacific_retail_2026.xlsx",
        "insured": "Pacific Retail Group Inc.",
        "broker_firm": "Aon",
        "broker_contact": "Paul Henderson",
        "eff_date": datetime.date(2026, 4, 1),
        "exp_date": datetime.date(2027, 4, 1),
        "coverage": "All-Risk Property incl. BI",
        "tiv": 120_000_000,
        "locations": 47,
        "our_line_pct": 1.00,
        "our_premium": 504_000,
        "technical_premium": 540_000,
        "decision": "Bound",
        "decision_date": datetime.date(2026, 3, 15),
        "underwriter": "James Chen",
        "occupancy": "Retail",
        "construction": "Non-Combustible (ISO Class 4)",
        "territory": "Southeast US",
        "base_rate": 4.50,
        "territory_factor": 1.10,
        "occupancy_factor": 0.95,
        "construction_factor": 0.90,
        "protection_factor": 0.85,
        "experience_mod": 0.92,
        "schedule_mod": 0.95,
        "uw_judgment": 0.98,
        "package_credit": 0.97,
        "technical_rate": 4.50,
        "quoted_rate": 4.20,
        "cat_wind": 0.45,
        "cat_eq": 0.05,
        "cat_flood": 0.10,
        "expense_ratio": 0.32,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 0, 0, 0.00),
            (2022, 0, 0, 0.00),
            (2023, 1, 125_000, 0.26),
            (2024, 0, 0, 0.00),
            (2025, 0, 0, 0.00),
        ],
        "large_losses": "Single $125K kitchen fire in 2023 -- contained by sprinkler system; no spread.",
        "loss_dev_notes": "All claims fully developed. No adverse development expected.",
        "uw_notes": (
            "Strong account, long-standing Aon relationship. 47 retail locations, "
            "predominantly SE US. 85% sprinklered, modern construction post-2005. "
            "Clean loss history -- single $125K kitchen fire contained by sprinklers.\n\n"
            "TIV $120M, quoting at $4.20/thousand which is slightly below technical "
            "$4.50 but justified by loss experience, account size, and strategic Aon "
            "relationship. Similar to Eastwood Retail (2024) written at $4.50 and "
            "Westfield Shopping (2023) at $4.80. Pacific Retail is better quality "
            "construction.\n\n"
            "Recommend bind at $504K. No fac needed -- well within net retention. "
            "Simon approved deviation."
        ),
    },
    # 2 - Acme Industrial
    {
        "filename": "RS_acme_industrial_2026.xlsx",
        "insured": "Acme Industrial Corp.",
        "broker_firm": "Willis",
        "broker_contact": "Claire Donovan",
        "eff_date": datetime.date(2026, 5, 1),
        "exp_date": datetime.date(2027, 5, 1),
        "coverage": "All-Risk Property incl. BI & Equipment Breakdown",
        "tiv": 45_000_000,
        "locations": 3,
        "our_line_pct": 1.00,
        "our_premium": 247_500,
        "technical_premium": 292_500,
        "decision": "Bound",
        "decision_date": datetime.date(2026, 4, 2),
        "underwriter": "James Chen",
        "occupancy": "Heavy Manufacturing",
        "construction": "Non-Combustible (ISO Class 4)",
        "territory": "Midwest US",
        "base_rate": 6.50,
        "territory_factor": 0.95,
        "occupancy_factor": 1.15,
        "construction_factor": 0.90,
        "protection_factor": 1.05,
        "experience_mod": 0.98,
        "schedule_mod": 1.00,
        "uw_judgment": 1.00,
        "package_credit": 0.97,
        "technical_rate": 6.50,
        "quoted_rate": 5.50,
        "cat_wind": 0.10,
        "cat_eq": 0.05,
        "cat_flood": 0.08,
        "expense_ratio": 0.30,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 0, 0, 0.00),
            (2022, 0, 0, 0.00),
            (2023, 1, 200_000, 0.83),
            (2024, 1, 85_000, 0.35),
            (2025, 0, 0, 0.00),
        ],
        "large_losses": "$200K water main burst (2023) -- inventory loss in Building 2.\n$85K electrical fire (2024) -- shop floor, caught early by night watchman.",
        "loss_dev_notes": "Both claims closed and fully developed. No subrogation recovery expected on water main.",
        "uw_notes": (
            "Acme Industrial -- metalworking and assembly, 3 facilities in OH/PA. "
            "TIV $45M. Main plant is non-combustible but Building 3 (warehouse, 1978) "
            "is joisted masonry and NOT sprinklered -- this is the key risk.\n\n"
            "Two claims in 5 years: $85K electrical fire in the shop (2024, caught early "
            "by night watchman) and $200K water main burst (2023, inventory loss). "
            "Technical rate $6.50/thousand, broker pushing for $5.50. Deviation of 15.4% "
            "exceeds my Level 2 authority. Referred to Simon.\n\n"
            "Comparable: Timber Creek Manufacturing (2024) bound at $5.80 with similar "
            "profile but fully sprinklered. The unsprinklered warehouse is the sticking "
            "point. If Willis agrees to $5.50 with a $50K deductible increase and "
            "sprinkler installation warranty within 18 months, I'd recommend binding. "
            "Simon approved with sprinkler warranty condition."
        ),
    },
    # 3 - Northern Manufacturing
    {
        "filename": "RS_northern_mfg_2026.xlsx",
        "insured": "Northern Manufacturing Ltd.",
        "broker_firm": "Marsh",
        "broker_contact": "James Whitfield",
        "eff_date": datetime.date(2026, 6, 1),
        "exp_date": datetime.date(2027, 6, 1),
        "coverage": "All-Risk Property",
        "tiv": 28_000_000,
        "locations": 1,
        "our_line_pct": 1.00,
        "our_premium": 123_200,
        "technical_premium": 123_200,
        "decision": "Bound",
        "decision_date": datetime.date(2026, 5, 10),
        "underwriter": "James Chen",
        "occupancy": "Light Manufacturing",
        "construction": "Fire Resistive (ISO Class 6)",
        "territory": "Midwest US",
        "base_rate": 4.40,
        "territory_factor": 0.95,
        "occupancy_factor": 1.00,
        "construction_factor": 0.85,
        "protection_factor": 0.85,
        "experience_mod": 0.90,
        "schedule_mod": 1.00,
        "uw_judgment": 1.00,
        "package_credit": 1.00,
        "technical_rate": 4.40,
        "quoted_rate": 4.40,
        "cat_wind": 0.08,
        "cat_eq": 0.03,
        "cat_flood": 0.05,
        "expense_ratio": 0.30,
        "profit_margin": 0.05,
        "loss_years": [
            (2023, 0, 0, 0.00),
            (2024, 0, 0, 0.00),
            (2025, 0, 0, 0.00),
        ],
        "large_losses": "None.",
        "loss_dev_notes": "No losses reported in 3 years on our book.",
        "uw_notes": (
            "Straightforward renewal. Northern Manufacturing -- single facility, light "
            "manufacturing, fire resistive construction (2015), fully sprinklered. TIV "
            "$28M. Zero claims in 3 years on our book. Technical rate matches quoted at "
            "$4.40/thousand. No deviation, no referral needed.\n\n"
            "Marsh account, James Whitfield always sends clean submissions. This is "
            "the kind of risk we want more of -- modern construction, single location, "
            "clean history, experienced broker. Renew at $123.2K flat. No changes to terms."
        ),
    },
    # 4 - Heritage Hotels
    {
        "filename": "RS_heritage_hotels_2026.xlsx",
        "insured": "Heritage Hotels International",
        "broker_firm": "Lockton",
        "broker_contact": "Emily Richards",
        "eff_date": datetime.date(2026, 7, 1),
        "exp_date": datetime.date(2027, 7, 1),
        "coverage": "All-Risk Property incl. BI, Windstorm",
        "tiv": 200_000_000,
        "locations": 12,
        "our_line_pct": 0.00,
        "our_premium": 0,
        "technical_premium": 1_700_000,
        "decision": "Declined",
        "decision_date": datetime.date(2026, 3, 28),
        "underwriter": "James Chen",
        "occupancy": "Hotel / Hospitality",
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "territory": "Florida",
        "base_rate": 8.50,
        "territory_factor": 1.35,
        "occupancy_factor": 1.10,
        "construction_factor": 0.95,
        "protection_factor": 0.90,
        "experience_mod": 1.00,
        "schedule_mod": 1.00,
        "uw_judgment": 1.00,
        "package_credit": 1.00,
        "technical_rate": 8.50,
        "quoted_rate": 5.50,
        "cat_wind": 3.80,
        "cat_eq": 0.10,
        "cat_flood": 0.60,
        "expense_ratio": 0.33,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 0, 0, 0.00),
            (2022, 1, 350_000, 0.32),
            (2023, 2, 1_200_000, 1.10),
            (2024, 0, 0, 0.00),
            (2025, 1, 180_000, 0.16),
        ],
        "large_losses": "$800K wind/water damage from Tropical Storm (2023) at Turks & Caicos property.\n$400K flood loss at Ft. Lauderdale location (2023).",
        "loss_dev_notes": "2023 tropical storm claim still developing -- $50K reserve remaining. Flood claim closed.",
        "uw_notes": (
            "Heritage Hotels -- 12 luxury properties, major red flag: 8 of 12 are "
            "coastal FL and Caribbean. TIV $200M with approximately $140M in Tier 1 "
            "wind zones. Running the CAT numbers, PML at 250yr return is $85M -- well "
            "above our net tolerance.\n\n"
            "CAT load alone pushes minimum viable rate to $8.50/thousand. Lockton is "
            "looking for $5.50 which is dramatically below what the risk requires. Even "
            "with fac, the net economics don't work -- fac market is extremely tight for "
            "Caribbean hospitality post-Ian.\n\n"
            "Similar to Beachfront Resorts (2023) and Coastal Luxury Properties (2024), "
            "both declined for identical reasons. The non-coastal 4 properties are fine "
            "risks but we can't cherry-pick the schedule. If they restructured to exclude "
            "Caribbean locations, we could revisit. As submitted, this is a clear decline. "
            "Emily at Lockton understands our position."
        ),
    },
    # 5 - Western Distribution
    {
        "filename": "RS_western_dist_2026.xlsx",
        "insured": "Western Distribution Services LLC",
        "broker_firm": "Howden",
        "broker_contact": "Michael Cross",
        "eff_date": datetime.date(2026, 8, 1),
        "exp_date": datetime.date(2027, 8, 1),
        "coverage": "All-Risk Property incl. BI",
        "tiv": 65_000_000,
        "locations": 4,
        "our_line_pct": 1.00,
        "our_premium": 468_000,
        "technical_premium": 468_000,
        "decision": "Bound",
        "decision_date": datetime.date(2026, 7, 15),
        "underwriter": "James Chen",
        "occupancy": "Warehouse / Distribution",
        "construction": "Non-Combustible (ISO Class 4)",
        "territory": "Pacific Northwest US",
        "base_rate": 7.20,
        "territory_factor": 0.90,
        "occupancy_factor": 1.05,
        "construction_factor": 0.90,
        "protection_factor": 0.90,
        "experience_mod": 1.15,
        "schedule_mod": 1.00,
        "uw_judgment": 1.05,
        "package_credit": 0.97,
        "technical_rate": 7.20,
        "quoted_rate": 7.20,
        "cat_wind": 0.12,
        "cat_eq": 0.15,
        "cat_flood": 0.08,
        "expense_ratio": 0.31,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 0, 0, 0.00),
            (2022, 0, 0, 0.00),
            (2023, 0, 0, 0.00),
            (2024, 2, 222_000, 0.57),
            (2025, 2, 130_000, 0.33),
        ],
        "large_losses": "$180K forklift battery fire (2024) -- warehouse section.\n$95K roof collapse from snow load (2025).\n$42K break-in (2025).\n$35K HVAC water leak (2024).",
        "loss_dev_notes": "Forklift fire claim closed. Snow collapse has $10K reserve remaining. All others closed.",
        "uw_notes": (
            "Western Distribution renewal -- need to address deteriorating loss "
            "experience. Expiring at $390K ($6.00/thousand on $65M TIV). Since binding "
            "18 months ago: forklift battery fire ($180K), roof collapse from snow "
            "($95K), break-in ($42K), and HVAC water leak ($35K). Four claims totalling "
            "$352K. 5-year loss ratio now 47% vs 28% at inception.\n\n"
            "Recommending renewal at $7.20/thousand ($468K) -- 20% rate increase. This "
            "is justified and within technical. Howden will push back but Michael Cross "
            "is pragmatic and knows the loss record.\n\n"
            "Compare to Dominion Storage (2025) where we got 18% increase after 3 claims. "
            "If they walk, the loss trend was wrong anyway. Retaining at adequate pricing "
            "is better than winning at inadequate pricing."
        ),
    },
    # 6 - Cascade Technology
    {
        "filename": "RS_cascade_tech_2026.xlsx",
        "insured": "Cascade Technology Campus LLC",
        "broker_firm": "Gallagher",
        "broker_contact": "Rachel Foster",
        "eff_date": datetime.date(2026, 5, 1),
        "exp_date": datetime.date(2027, 5, 1),
        "coverage": "All-Risk Property incl. BI & Equipment Breakdown",
        "tiv": 180_000_000,
        "locations": 2,
        "our_line_pct": 1.00,
        "our_premium": 630_000,
        "technical_premium": 684_000,
        "decision": "Bound",
        "decision_date": datetime.date(2026, 4, 10),
        "underwriter": "James Chen",
        "occupancy": "Data Centre",
        "construction": "Fire Resistive (ISO Class 6)",
        "territory": "Pacific Northwest US",
        "base_rate": 3.80,
        "territory_factor": 0.90,
        "occupancy_factor": 0.85,
        "construction_factor": 0.80,
        "protection_factor": 0.80,
        "experience_mod": 0.88,
        "schedule_mod": 0.95,
        "uw_judgment": 0.95,
        "package_credit": 0.97,
        "technical_rate": 3.80,
        "quoted_rate": 3.50,
        "cat_wind": 0.05,
        "cat_eq": 0.25,
        "cat_flood": 0.05,
        "expense_ratio": 0.30,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 0, 0, 0.00),
            (2022, 0, 0, 0.00),
            (2023, 0, 0, 0.00),
            (2024, 0, 0, 0.00),
            (2025, 0, 0, 0.00),
        ],
        "large_losses": "None.",
        "loss_dev_notes": "No losses. Facilities operational since 2019/2020.",
        "uw_notes": (
            "Exceptional risk quality. Cascade Technology Campus -- 2 data centres in "
            "Seattle metro, TIV $180M. State-of-the-art facilities: fire resistive "
            "steel/concrete, 2019-2020 build, N+1 redundancy on all critical systems, "
            "FM Global-approved fire suppression (clean agent + pre-action sprinklers), "
            "24/7 onsite security and monitoring. No loss history.\n\n"
            "Low natural hazard exposure -- Seattle is minimal wind, moderate seismic "
            "but building is base-isolated. Quoting at $3.50/thousand vs technical $3.80. "
            "The below-technical rate is justified: this is arguably the best-protected "
            "property class we underwrite.\n\n"
            "Similar to Quantum Data Systems (2024) bound at $3.60 and Richmond Data "
            "Centres (2025) at $3.40. Gallagher has 4 more data centre accounts they "
            "want to place with us if we perform well here. Strategic account. Rachel "
            "Foster is excellent to work with.\n\n"
            "Recommend bind at $630K. Fac placed for excess layer above $25M net."
        ),
    },
    # 7 - Gulf Coast Energy
    {
        "filename": "RS_gulf_coast_energy_2026.xlsx",
        "insured": "Gulf Coast Energy Storage Inc.",
        "broker_firm": "McGill",
        "broker_contact": "Thomas Keane",
        "eff_date": datetime.date(2026, 6, 1),
        "exp_date": datetime.date(2027, 6, 1),
        "coverage": "All-Risk Property incl. BI & Windstorm",
        "tiv": 95_000_000,
        "locations": 5,
        "our_line_pct": 1.00,
        "our_premium": 712_500,
        "technical_premium": 760_000,
        "decision": "Referred",
        "decision_date": datetime.date(2026, 5, 5),
        "underwriter": "James Chen",
        "occupancy": "Chemical Processing",
        "construction": "Non-Combustible (ISO Class 4)",
        "territory": "Gulf Coast US",
        "base_rate": 8.00,
        "territory_factor": 1.30,
        "occupancy_factor": 1.10,
        "construction_factor": 0.90,
        "protection_factor": 0.90,
        "experience_mod": 0.98,
        "schedule_mod": 1.00,
        "uw_judgment": 1.00,
        "package_credit": 0.97,
        "technical_rate": 8.00,
        "quoted_rate": 7.50,
        "cat_wind": 2.80,
        "cat_eq": 0.05,
        "cat_flood": 0.40,
        "expense_ratio": 0.32,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 0, 0, 0.00),
            (2022, 0, 0, 0.00),
            (2023, 0, 0, 0.00),
            (2024, 1, 145_000, 0.19),
            (2025, 0, 0, 0.00),
        ],
        "large_losses": "$145K tropical storm roof/piping damage (2024) at Galveston facility. $25K reserved.",
        "loss_dev_notes": "2024 storm claim: $120K paid, $25K reserve for remaining repairs. Expected to close Q2 2026.",
        "uw_notes": (
            "Gulf Coast Energy Storage -- 5 facilities along TX coast, TIV $95M. Energy "
            "storage is a growth class for us but this one has genuine CAT complexity. "
            "Houston, Galveston, Corpus Christi, Beaumont, Port Arthur -- all in wind "
            "Tier 1-2 zones.\n\n"
            "One historical claim: $145K from tropical storm roof/piping damage (2024), "
            "currently $25K still reserved. The underlying risk quality is decent -- "
            "non-combustible, 2016-2020 construction, sprinklered, proper containment. "
            "But aggregate wind PML is $42M at 100yr return.\n\n"
            "Technical rate $8.00/thousand, broker (Thomas Keane, McGill) targeting $7.50. "
            "Deviation is only 6.25% but the TIV and CAT exposure push this above my "
            "Level 2 authority. Referring to Simon.\n\n"
            "My recommendation: bind with named storm deductible of 3% per location and "
            "aggregate CAT deductible of $2M. This brings net CAT exposure to manageable "
            "level. If Simon agrees to terms, the energy sector expertise this builds is "
            "worth the allocation."
        ),
    },
    # 8 - Midwest Ag
    {
        "filename": "RS_midwest_ag_2026.xlsx",
        "insured": "Midwest Agricultural Cooperative",
        "broker_firm": "Miller Insurance",
        "broker_contact": "Charlotte Webb",
        "eff_date": datetime.date(2026, 4, 1),
        "exp_date": datetime.date(2027, 4, 1),
        "coverage": "All-Risk Property",
        "tiv": 35_000_000,
        "locations": 22,
        "our_line_pct": 1.00,
        "our_premium": 210_000,
        "technical_premium": 217_000,
        "decision": "Bound",
        "decision_date": datetime.date(2026, 3, 20),
        "underwriter": "Rebecca Hartley",
        "occupancy": "Agricultural",
        "construction": "Heavy Timber (ISO Class 2)",
        "territory": "Midwest US",
        "base_rate": 6.20,
        "territory_factor": 0.95,
        "occupancy_factor": 1.10,
        "construction_factor": 1.10,
        "protection_factor": 1.10,
        "experience_mod": 0.92,
        "schedule_mod": 1.00,
        "uw_judgment": 0.98,
        "package_credit": 0.97,
        "technical_rate": 6.20,
        "quoted_rate": 6.00,
        "cat_wind": 0.15,
        "cat_eq": 0.02,
        "cat_flood": 0.10,
        "expense_ratio": 0.30,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 1, 78_000, 0.38),
            (2022, 0, 0, 0.00),
            (2023, 1, 45_000, 0.22),
            (2024, 1, 32_000, 0.15),
            (2025, 0, 0, 0.00),
        ],
        "large_losses": "No individual loss exceeds $100K.",
        "loss_dev_notes": "All claims closed. Predominantly weather-related attritional losses.",
        "uw_notes": (
            "Midwest Agricultural Coop -- 22 grain storage and processing facilities "
            "across MN, IA, WI, ND. TIV $35M, lots of small locations ($800K-$3M each). "
            "Heavy timber construction, mostly unsprinklered -- typical for ag sector.\n\n"
            "Three claims in 5 years: hail roof damage ($78K), lightning/electrical surge "
            "($45K), wind damage ($32K). All weather-related, all modest. Total loss ratio "
            "25% which is respectable for agricultural.\n\n"
            "Technical rate $6.20/thousand, binding at $6.00 -- within tolerance. The "
            "challenge is the number of locations (22) makes this operationally complex "
            "for a $35M TIV. But the geographic spread actually reduces aggregation risk "
            "-- no more than 4 locations in any single county.\n\n"
            "Charlotte Webb at Miller Insurance has placed 6 similar ag accounts with us "
            "and all have performed well. Recommend bind at $210K."
        ),
    },
    # 9 - Atlantic Seafood
    {
        "filename": "RS_atlantic_seafood_2026.xlsx",
        "insured": "Atlantic Seafood Processing Inc.",
        "broker_firm": "BMS Group",
        "broker_contact": "Andrew Patel",
        "eff_date": datetime.date(2026, 7, 1),
        "exp_date": datetime.date(2027, 7, 1),
        "coverage": "All-Risk Property incl. BI & Equipment Breakdown",
        "tiv": 50_000_000,
        "locations": 4,
        "our_line_pct": 1.00,
        "our_premium": 375_000,
        "technical_premium": 350_000,
        "decision": "Bound",
        "decision_date": datetime.date(2026, 6, 15),
        "underwriter": "James Chen",
        "occupancy": "Food Processing",
        "construction": "Non-Combustible (ISO Class 4)",
        "territory": "Northeast US",
        "base_rate": 7.00,
        "territory_factor": 1.00,
        "occupancy_factor": 1.15,
        "construction_factor": 0.90,
        "protection_factor": 0.90,
        "experience_mod": 1.10,
        "schedule_mod": 1.00,
        "uw_judgment": 1.05,
        "package_credit": 0.97,
        "technical_rate": 7.00,
        "quoted_rate": 7.50,
        "cat_wind": 0.30,
        "cat_eq": 0.05,
        "cat_flood": 0.15,
        "expense_ratio": 0.31,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 0, 0, 0.00),
            (2022, 0, 0, 0.00),
            (2023, 0, 0, 0.00),
            (2024, 0, 0, 0.00),
            (2025, 1, 2_100_000, 5.60),
        ],
        "large_losses": "$2.1M refrigeration system failure (2025) -- catastrophic compressor failure, complete cold storage inventory loss ($1.8M) plus equipment and BI. $150K reserved.",
        "loss_dev_notes": "2025 large loss: $1.95M paid, $150K reserve for final BI settlement. Expected to close Q3 2026.",
        "uw_notes": (
            "Atlantic Seafood renewal -- this is the one with the major refrigeration "
            "loss. TIV $50M across 4 processing plants in coastal MA. The $2.1M "
            "refrigeration system failure last year is the elephant in the room -- "
            "complete cold storage inventory loss ($1.8M) plus equipment replacement and "
            "BI. Still $150K reserved.\n\n"
            "Prior to that loss, account was clean. The loss was a mechanical failure, "
            "not a coverage or maintenance issue -- their equipment was maintained per "
            "manufacturer spec, just a catastrophic compressor failure that cascaded. "
            "But it moved the 5-year loss ratio to 58%.\n\n"
            "Technical rate is $7.00/thousand but I'm pricing at $7.50 -- ABOVE "
            "technical by 7% -- to account for the frequency risk in food processing "
            "and the recency of the large loss. Premium $375K. BMS Group (Andrew Patel) "
            "understands -- they expected a rate increase.\n\n"
            "We've required a backup refrigeration system warranty and quarterly "
            "maintenance certification. If they have another large loss, we non-renew. "
            "This is a well-run operation that had bad luck."
        ),
    },
    # 10 - Summit Office
    {
        "filename": "RS_summit_office_2026.xlsx",
        "insured": "Summit Office Partners LP",
        "broker_firm": "Ed Broking",
        "broker_contact": "Laura Simpson",
        "eff_date": datetime.date(2026, 6, 1),
        "exp_date": datetime.date(2027, 6, 1),
        "coverage": "All-Risk Property",
        "tiv": 75_000_000,
        "locations": 6,
        "our_line_pct": 0.25,
        "our_premium": 187_500,
        "technical_premium": 187_500,
        "decision": "Bound",
        "decision_date": datetime.date(2026, 5, 20),
        "underwriter": "Rebecca Hartley",
        "occupancy": "Commercial Office",
        "construction": "Fire Resistive (ISO Class 6)",
        "territory": "Northeast US",
        "base_rate": 2.50,
        "territory_factor": 1.00,
        "occupancy_factor": 0.85,
        "construction_factor": 0.80,
        "protection_factor": 0.85,
        "experience_mod": 0.95,
        "schedule_mod": 1.00,
        "uw_judgment": 1.00,
        "package_credit": 1.00,
        "technical_rate": 2.50,
        "quoted_rate": 2.50,
        "cat_wind": 0.10,
        "cat_eq": 0.08,
        "cat_flood": 0.05,
        "expense_ratio": 0.30,
        "profit_margin": 0.05,
        "loss_years": [
            (2021, 0, 0, 0.00),
            (2022, 0, 0, 0.00),
            (2023, 1, 38_000, 0.20),
            (2024, 0, 0, 0.00),
            (2025, 0, 0, 0.00),
        ],
        "large_losses": "Single $38K pipe burst (2023) -- closed.",
        "loss_dev_notes": "All claims closed. Minimal loss activity.",
        "uw_notes": (
            "Summit Office Partners -- follow placement behind Chubb as lead at "
            "$2.50/thousand. 6 Class A office buildings in NY metro. TIV $75M, we're "
            "taking 25% ($18.75M line).\n\n"
            "This is vanilla: fire resistive construction, 2010-2018 vintage, fully "
            "sprinklered, central station alarms. One small claim ($38K pipe burst, "
            "closed). Chubb priced it well -- $2.50 is fair for Class A NY office.\n\n"
            "Our line is within authority. Ed Broking (Laura Simpson) handles the "
            "placement. Minimal work for steady premium. Follow terms as quoted, bind "
            "at $187.5K for our 25% line."
        ),
    },
]


# ---------------------------------------------------------------------------
# Templates for the 40 non-hero accounts
# ---------------------------------------------------------------------------

HISTORICAL_BOUND = [
    # 1 - Healthcare
    {
        "filename": "RS_brighton_health_2025.xlsx",
        "insured": "Brighton Health System",
        "occupancy": "Healthcare / Hospital",
        "territory": "Northeast US",
        "tiv": 150_000_000, "locations": 8,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "Brighton Health System -- 8 hospital and outpatient facilities across NJ "
            "and eastern PA. TIV $150M. All fire resistive, fully sprinklered with "
            "redundant fire suppression in surgical suites and data rooms. Healthcare "
            "is an occupancy we understand well.\n\n"
            "Clean loss history: one $95K water damage claim (burst pipe in radiology "
            "wing, 2023). 5-year loss ratio is 8%. Technical rate $3.20/thousand, "
            "quoting at $3.00 -- 6.25% deviation justified by the excellent loss "
            "experience and the fact that Aon is bringing us their broader healthcare "
            "portfolio.\n\n"
            "Equipment breakdown coverage is critical here -- medical imaging equipment "
            "values alone are $18M. We've attached EQ sub-limit of $15M which is "
            "appropriate for the region. BI waiting period of 24 hours is standard.\n\n"
            "Bind at $450K. Paul Henderson at Aon has been responsive throughout. "
            "Good account for the book."
        ),
    },
    # 2 - Pharma
    {
        "filename": "RS_novus_pharma_2025.xlsx",
        "insured": "Novus Pharmaceutical Inc.",
        "occupancy": "Pharmaceutical",
        "territory": "Mid-Atlantic US",
        "tiv": 85_000_000, "locations": 3,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "Novus Pharmaceutical -- 3 facilities in NJ (2 manufacturing, 1 R&D lab). "
            "TIV $85M. Pharma is a higher-hazard occupancy but Novus is a well-managed "
            "operation with FDA cGMP compliance and strong loss control programme.\n\n"
            "Stock values are significant ($22M in finished pharmaceutical products) "
            "and temperature-sensitive -- they have redundant cold chain infrastructure. "
            "One claim in 5 years: $160K HVAC failure in clean room (2022), fully "
            "remediated. 5-year loss ratio 12%.\n\n"
            "Technical rate $4.80/thousand, quoting at $4.50 -- 6.25% below technical. "
            "The deviation is acceptable given loss experience and the quality of their "
            "risk management programme. Marsh placed this originally -- Sarah Price is "
            "the contact.\n\n"
            "Key coverage point: we've endorsed the policy with a contamination "
            "exclusion for the manufacturing facilities and added a product withdrawal "
            "sub-limit of $5M. Bind at $382.5K."
        ),
    },
    # 3 - Automotive
    {
        "filename": "RS_sterling_auto_2025.xlsx",
        "insured": "Sterling Automotive Group",
        "occupancy": "Automotive",
        "territory": "Midwest US",
        "tiv": 42_000_000, "locations": 6,
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "Sterling Automotive -- 6 car dealership locations across IL and IN. "
            "TIV $42M. Standard auto dealership risk with typical exposures: large "
            "open floor plans, paint booths in service areas (all NFPA-compliant), "
            "significant vehicle inventory exposed.\n\n"
            "Two claims in 5 years: $55K hail damage to vehicle inventory (2023) "
            "and $28K electrical fire in service bay (2024, sprinkler-controlled). "
            "Loss ratio 13% which is excellent for auto dealership.\n\n"
            "Technical rate $5.00/thousand. Binding at $4.80 -- slight discount for "
            "the multi-location programme and clean history. Willis account, Robert "
            "Chen handles. Vehicle inventory reporting endorsed quarterly.\n\n"
            "Bind at $201.6K. Straightforward risk."
        ),
    },
    # 4 - UK Real Estate
    {
        "filename": "RS_crown_properties_2025.xlsx",
        "insured": "Crown Properties (UK) Ltd.",
        "occupancy": "Real Estate Portfolio",
        "territory": "Central London",
        "tiv": 220_000_000, "locations": 15,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "Crown Properties -- prestigious portfolio of 15 commercial office "
            "buildings in central London (City, Canary Wharf, West End). TIV GBP 175M "
            "(approx $220M). All Grade A office space, fire resistive construction, "
            "fully sprinklered.\n\n"
            "This is a follow placement -- Hiscox leads at GBP 1.80/thousand, we're "
            "taking 20% line. Loss history is clean: single GBP 45K water escape claim "
            "in 2024. London commercial office is a bread-and-butter class for us.\n\n"
            "Terrorism cover included via Pool Re. Flood exposure is managed -- all "
            "properties are in Flood Zone 1 (low risk) per EA mapping. No subsidence "
            "history.\n\n"
            "Our line premium: GBP 63K ($79.5K equivalent). Ed Broking handles the "
            "placement -- Greg Thomson is the broker. Minimal CAT exposure. Good "
            "diversification for the London book."
        ),
    },
    # 5 - Self-Storage
    {
        "filename": "RS_securestore_2025.xlsx",
        "insured": "SecureStore Holdings Corp.",
        "occupancy": "Self-Storage",
        "territory": "Southeast US",
        "tiv": 28_000_000, "locations": 12,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "SecureStore -- 12 self-storage facilities across GA, SC, NC. TIV $28M. "
            "Non-combustible construction (metal buildings), limited fire load as "
            "building operator rather than stock owner. Customer contents are excluded "
            "from our policy -- we're insuring structures only.\n\n"
            "No claims in 3 years on our book. Self-storage is a favourable occupancy: "
            "low complexity, low loss frequency, simple operations. Technical rate "
            "$3.60/thousand, binding at $3.40 -- small deviation for programme size "
            "and clean experience.\n\n"
            "Key consideration: SE US wind exposure, but metal buildings are well-rated "
            "for wind (low-profile, no significant contents). Named storm deductible of "
            "2% per location applies.\n\n"
            "Bind at $95.2K. Howden account, Helen Taylor. Renews annually with minimal "
            "servicing -- ideal for portfolio."
        ),
    },
    # 6 - Electronics Assembly
    {
        "filename": "RS_precision_electronics_2024.xlsx",
        "insured": "Precision Electronics Inc.",
        "occupancy": "Electronics Assembly",
        "territory": "California",
        "tiv": 55_000_000, "locations": 2,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Bound", "year": 2024,
        "uw_notes": (
            "Precision Electronics -- 2 assembly and testing facilities in San Jose. "
            "TIV $55M. Clean room environments with state-of-the-art fire suppression "
            "(Novec 1230 clean agent systems). High-value stock -- finished electronic "
            "components valued at $15M.\n\n"
            "EQ is the key peril here. San Jose is Zone 4 seismic. Building is bolted "
            "steel frame on reinforced mat foundation -- engineered for seismic. We've "
            "run PML at $18M (250yr return). EQ deductible is 5% of TIV ($2.75M) which "
            "is standard for the zone.\n\n"
            "One claim: $72K water damage from fire suppression discharge (false alarm, "
            "2022). Loss ratio 8%. Technical rate $4.20/thousand, binding at $4.00 -- "
            "slight discount for loss experience. Gallagher account, Kevin O'Brien.\n\n"
            "Bind at $220K. Clean risk, well-managed. Stock reporting endorsed quarterly."
        ),
    },
    # 7 - Printing
    {
        "filename": "RS_national_print_2024.xlsx",
        "insured": "National Print & Publishing Group",
        "occupancy": "Printing / Publishing",
        "territory": "Northeast US",
        "tiv": 32_000_000, "locations": 2,
        "construction": "Joisted Masonry (ISO Class 3)",
        "decision": "Bound", "year": 2024,
        "uw_notes": (
            "National Print -- 2 printing facilities in CT. TIV $32M. Printing is "
            "a higher-hazard occupancy due to solvent use, paper stock, and heat-"
            "generating equipment. Both facilities are sprinklered with proper solvent "
            "storage (FM-approved cabinets, proper ventilation).\n\n"
            "Joisted masonry construction is the weaker point -- these are older "
            "buildings (1985, 1992) but well-maintained. Fire walls separate press "
            "rooms from paper storage.\n\n"
            "Two claims in 5 years: $42K ink spill/cleanup (2022) and $65K press "
            "roller fire (2023, extinguished by sprinklers before spread). Loss ratio "
            "21%. Technical rate $5.80/thousand, binding at $5.60.\n\n"
            "Marsh account, Tom Bradley. This is a mature account -- we've had it for "
            "4 years. Consistent performer. Bind at $179.2K."
        ),
    },
    # 8 - UK Mixed-Use
    {
        "filename": "RS_meridian_mixed_2024.xlsx",
        "insured": "Meridian Mixed-Use Developments Ltd.",
        "occupancy": "Mixed-Use Commercial",
        "territory": "Greater Manchester",
        "tiv": 65_000_000, "locations": 4,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "Bound", "year": 2024,
        "uw_notes": (
            "Meridian Mixed-Use -- 4 modern mixed-use developments in Greater "
            "Manchester. Retail ground floor, commercial office above. TIV GBP 52M "
            "(approx $65M). All post-2015 construction, fire resistive, fully "
            "sprinklered, compartmented per UK Building Regs.\n\n"
            "No claims on our book (bound 2023). Manchester is a growing commercial "
            "market with solid occupancy rates. The tenant mix is stable -- no high-"
            "hazard retail occupancies (no restaurants/food prep in the schedule).\n\n"
            "Technical rate GBP 2.80/thousand, binding at GBP 2.60. Miller Insurance "
            "placed this -- Ben Archer is responsive. Terrorism cover via Pool Re.\n\n"
            "Our premium GBP 135.2K ($169K equivalent). Good UK portfolio "
            "diversification. Renew at flat rate."
        ),
    },
    # 9 - Cold Storage
    {
        "filename": "RS_arctic_cold_2025.xlsx",
        "insured": "Arctic Cold Chain Logistics LLC",
        "occupancy": "Cold Storage / Refrigeration",
        "territory": "Mid-Atlantic US",
        "tiv": 40_000_000, "locations": 3,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "Arctic Cold Chain -- 3 temperature-controlled warehouse/distribution "
            "facilities in MD and VA. TIV $40M. Cold storage is a specialty class "
            "that requires careful underwriting due to ammonia refrigeration systems "
            "and high stock values.\n\n"
            "All three facilities use ammonia-based systems with proper PSM (Process "
            "Safety Management) programmes. EPA RMP-compliant. Backup diesel generators "
            "at all locations for power failure contingency. Stock values fluctuate "
            "seasonally -- peak at $12M in Q4.\n\n"
            "One claim: $88K ammonia leak/cleanup (2023). No injuries, proper "
            "containment worked. Loss ratio 14%. Technical rate $5.40/thousand, "
            "binding at $5.20.\n\n"
            "Willis account, Emily Watson. Quarterly stock reporting endorsed. "
            "Ammonia leak detection system warranty required. Bind at $208K."
        ),
    },
    # 10 - Educational
    {
        "filename": "RS_westbrook_academy_2024.xlsx",
        "insured": "Westbrook Academy Trust",
        "occupancy": "Educational",
        "territory": "South East England",
        "tiv": 75_000_000, "locations": 5,
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "decision": "Bound", "year": 2024,
        "uw_notes": (
            "Westbrook Academy Trust -- 5 independent school campuses across Surrey "
            "and Kent. TIV GBP 60M (approx $75M). Mix of historic (Grade II listed) "
            "and modern buildings. The listed buildings are the underwriting challenge "
            "-- higher reinstatement costs and limited fire protection options.\n\n"
            "The modern buildings (3 of 5 campuses, built 2000-2018) are fully "
            "sprinklered. The 2 historic campuses have detection only -- sprinklers "
            "not permitted by Historic England in the listed structures. We've applied "
            "a higher rate to the heritage buildings.\n\n"
            "Blended technical rate GBP 3.40/thousand. One claim: GBP 28K water damage "
            "at Cranbrook campus (2023). Loss ratio 2.5%. Low hazard occupancy overall.\n\n"
            "Lockton account, Jason Brooks. Bind at GBP 204K ($255K). Terrorism via "
            "Pool Re. Good account -- educational is a stable, low-loss class."
        ),
    },
    # 11 - Warehouse
    {
        "filename": "RS_continental_logistics_2025.xlsx",
        "insured": "Continental Logistics Partners",
        "occupancy": "Warehouse / Distribution",
        "territory": "Southwest US",
        "tiv": 90_000_000, "locations": 5,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "Continental Logistics -- 5 large distribution centres in AZ and NV. "
            "TIV $90M. Non-combustible tilt-up concrete construction, ESFR sprinklers "
            "throughout. Modern facilities (2017-2022 build). High-bay racking up to "
            "40ft -- ESFR design is critical for this storage arrangement.\n\n"
            "Two claims in 5 years: $35K forklift collision damage (2024) and $48K "
            "roof leak from monsoon (2023). Loss ratio 6%. Excellent for warehousing.\n\n"
            "Minimal CAT exposure -- AZ/NV is very low wind and flood. Some wildfire "
            "exposure at the Henderson, NV location but defensible space exceeds 200ft "
            "and local fire department is 3 miles away.\n\n"
            "Technical rate $3.80/thousand, binding at $3.60. Aon account, Lisa Chang. "
            "Bind at $324K. This is the type of warehouse risk we actively target -- "
            "modern construction, ESFR sprinklers, low CAT zone."
        ),
    },
    # 12 - Healthcare UK
    {
        "filename": "RS_harley_medical_2024.xlsx",
        "insured": "Harley Medical Group Plc",
        "occupancy": "Healthcare / Hospital",
        "territory": "Central London",
        "tiv": 45_000_000, "locations": 3,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "Bound", "year": 2024,
        "uw_notes": (
            "Harley Medical Group -- 3 private medical facilities in London (Harley "
            "Street, Kensington, Chelsea). TIV GBP 36M (approx $45M). High-quality "
            "fire resistive construction, all fully sprinklered with specialised "
            "suppression in operating theatres and MRI suites.\n\n"
            "Private healthcare in central London is a well-understood class. The "
            "facilities are impeccably maintained -- CQC-rated Outstanding. No claims "
            "in 4 years on our book.\n\n"
            "Technical rate GBP 2.40/thousand, binding at GBP 2.20. Small deviation "
            "justified by zero loss experience and quality of risk. Guy Carpenter "
            "placed the programme -- Neil Armstrong.\n\n"
            "Bind at GBP 79.2K ($99K). Terrorism via Pool Re. Low-maintenance "
            "account, excellent fit for the London healthcare book."
        ),
    },
    # 13 - Hotel
    {
        "filename": "RS_parkview_hotels_2025.xlsx",
        "insured": "Parkview Hotels & Resorts Inc.",
        "occupancy": "Hotel / Hospitality",
        "territory": "Mountain West US",
        "tiv": 55_000_000, "locations": 4,
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "Parkview Hotels -- 4 mid-range hotel properties in CO and UT. TIV $55M. "
            "Mountain resort locations -- ski season is peak occupancy. Hotels are "
            "1995-2010 construction, masonry non-combustible, all sprinklered.\n\n"
            "Hospitality is a moderate-hazard class due to cooking, laundry, and "
            "24/7 occupancy. Parkview has a solid risk management programme including "
            "commercial kitchen suppression systems (Ansul) and guest room smoke "
            "detection.\n\n"
            "One claim: $110K kitchen fire at Vail property (2024) -- Ansul system "
            "activated and contained to kitchen. No guest injuries. Loss ratio 13%.\n\n"
            "Technical rate $4.60/thousand, binding at $4.40. Wildfire is a "
            "consideration for the CO properties -- all are in WUI (wildland-urban "
            "interface) but defensible space is maintained and local fire response "
            "is excellent.\n\n"
            "Lockton account, Amanda Foster. Bind at $242K."
        ),
    },
    # 14 - Light Manufacturing
    {
        "filename": "RS_apex_precision_2025.xlsx",
        "insured": "Apex Precision Components Ltd.",
        "occupancy": "Light Manufacturing",
        "territory": "Midlands UK",
        "tiv": 22_000_000, "locations": 1,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Bound", "year": 2025,
        "uw_notes": (
            "Apex Precision -- single facility CNC machining operation in Birmingham. "
            "TIV GBP 17.5M (approx $22M). Non-combustible portal frame construction "
            "(2012), fully sprinklered. Manufactures precision aerospace components -- "
            "high-value stock but well-protected.\n\n"
            "Light manufacturing with CNC equipment is a well-understood risk. Metal "
            "cutting generates swarf but Apex has proper extraction and coolant systems. "
            "No flammable liquids in significant quantities.\n\n"
            "Zero claims in 5 years on our book. Technical rate GBP 3.80/thousand, "
            "binding at GBP 3.60. Deviation is minimal and justified by loss record.\n\n"
            "Miller Insurance placed this, Lucy Ward. Bind at GBP 63K ($79K). Simple "
            "risk, steady premium. This is the kind of UK manufacturing account we "
            "want to build the book around."
        ),
    },
    # 15 - Retail UK
    {
        "filename": "RS_highstreet_retail_2024.xlsx",
        "insured": "Highstreet Retail Group Plc",
        "occupancy": "Retail",
        "territory": "North West England",
        "tiv": 38_000_000, "locations": 10,
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "decision": "Bound", "year": 2024,
        "uw_notes": (
            "Highstreet Retail Group -- 10 department store locations across North "
            "West England. TIV GBP 30M (approx $38M). Masonry non-combustible "
            "construction, mix of Victorian-era and 1960s buildings, all upgraded "
            "with modern fire detection and 7 of 10 sprinklered.\n\n"
            "Retail is a moderate hazard class. Highstreet has good housekeeping "
            "standards and staff fire training. The 3 unsprinklered locations are "
            "smaller (sub-GBP 2M TIV each) with detection and 24/7 monitoring.\n\n"
            "One claim: GBP 65K arson attempt at Manchester store (2023) -- fire "
            "contained by sprinklers. Loss ratio 11%. Technical rate GBP 4.20/"
            "thousand, binding at GBP 4.00.\n\n"
            "Howden account, Ian Russell. Bind at GBP 120K ($150K). Arson is a "
            "concern for high-street retail but Highstreet has good security "
            "including CCTV and roller shutters."
        ),
    },
]

HISTORICAL_DECLINED = [
    # 1
    {
        "filename": "RS_coastal_marina_2025.xlsx",
        "insured": "Coastal Marina & Resort Group",
        "occupancy": "Hotel / Hospitality",
        "territory": "Florida",
        "tiv": 180_000_000, "locations": 6,
        "construction": "Frame (ISO Class 1)",
        "decision": "Declined", "year": 2025,
        "uw_notes": (
            "Coastal Marina -- 6 waterfront resort properties in FL Keys and SW FL. "
            "TIV $180M. DECLINE.\n\n"
            "Multiple red flags: (1) Frame construction at 4 of 6 locations -- ISO "
            "Class 1 in a Tier 1 wind zone is essentially uninsurable at any "
            "reasonable rate. (2) 3 locations have no sprinklers. (3) Flood zone AE "
            "for 5 of 6 properties. (4) Loss history includes $2.4M Hurricane Irma "
            "claim (2017) and $800K tropical storm damage (2022).\n\n"
            "PML at 100yr return is $95M. The CAT load alone would require a rate "
            "of $12.00/thousand minimum. Broker (Marsh, Sarah Price) is seeking "
            "$6.00/thousand which is completely unrealistic.\n\n"
            "We've declined similar FL Keys hospitality three times in the past 2 "
            "years. Frame construction on the water in the Keys is not a class we "
            "can underwrite. Clear decline -- communicated to Sarah."
        ),
    },
    # 2
    {
        "filename": "RS_legacy_textiles_2025.xlsx",
        "insured": "Legacy Textiles Inc.",
        "occupancy": "Heavy Manufacturing",
        "territory": "Southeast US",
        "tiv": 25_000_000, "locations": 2,
        "construction": "Frame (ISO Class 1)",
        "decision": "Declined", "year": 2025,
        "uw_notes": (
            "Legacy Textiles -- 2 textile mills in NC, both 1960s-era frame "
            "construction. TIV $25M. DECLINE.\n\n"
            "Textile manufacturing in frame buildings is a severely hazardous "
            "occupancy. Lint accumulation, combustible raw materials, heat-"
            "generating equipment. Neither facility is sprinklered. One location "
            "has no fire detection at all.\n\n"
            "Loss history is terrible: 4 claims in 5 years totalling $1.8M. "
            "5-year loss ratio 180%. Latest claim was a $900K fire that destroyed "
            "the dyeing room (2024). Prior to that: $450K warehouse fire (2023), "
            "$280K equipment fire (2022), $170K stock loss from water damage (2021).\n\n"
            "No amount of rate adequacy fixes the fundamental risk quality issue "
            "here. The facilities need substantial investment in fire protection "
            "before we could consider. Willis brought this and I've explained our "
            "position. Clear decline."
        ),
    },
    # 3
    {
        "filename": "RS_phoenix_recycling_2025.xlsx",
        "insured": "Phoenix Recycling Services LLC",
        "occupancy": "Heavy Manufacturing",
        "territory": "Southwest US",
        "tiv": 18_000_000, "locations": 3,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Declined", "year": 2025,
        "uw_notes": (
            "Phoenix Recycling -- 3 waste processing and recycling facilities in "
            "AZ. TIV $18M. DECLINE.\n\n"
            "Recycling operations are among the highest-hazard occupancies we see. "
            "These facilities process mixed waste including metals, plastics, and "
            "paper. Spontaneous combustion risk is significant -- lithium batteries "
            "in the waste stream are a growing concern industry-wide.\n\n"
            "Two fires in 2 years: $320K yard fire from lithium battery ignition "
            "(2024) and $185K baler fire (2023). Neither sprinklered -- open-air "
            "yards are essentially unprotectable.\n\n"
            "Even at $12.00/thousand ($216K premium), the expected loss exceeds "
            "premium. The business model doesn't support the rate we'd need. "
            "McGill (Fiona Campbell) understands this is outside our appetite. "
            "Decline."
        ),
    },
    # 4
    {
        "filename": "RS_blackwood_timber_2024.xlsx",
        "insured": "Blackwood Timber Products Ltd.",
        "occupancy": "Heavy Manufacturing",
        "territory": "Scotland",
        "tiv": 15_000_000, "locations": 2,
        "construction": "Heavy Timber (ISO Class 2)",
        "decision": "Declined", "year": 2024,
        "uw_notes": (
            "Blackwood Timber -- 2 sawmill and timber processing facilities in "
            "the Scottish Highlands. TIV GBP 12M (approx $15M). DECLINE.\n\n"
            "Sawmill operations in heavy timber construction -- the building IS the "
            "fuel. Massive dust generation, kilns operating at 70C+, and limited "
            "fire service response (nearest station 22 minutes). One facility has "
            "basic detection only, no suppression of any kind.\n\n"
            "Loss history: GBP 400K kiln fire (2022) and GBP 150K dust explosion "
            "(2023). Combined loss ratio 220%. This is a risk profile we simply "
            "cannot accommodate.\n\n"
            "The remote location compounds every other issue. If there's a fire, "
            "it will be a total loss before the fire service arrives. Miller "
            "Insurance (Ben Archer) brought this -- I've explained that sawmill "
            "in remote Scotland with this loss record is outside our appetite "
            "entirely."
        ),
    },
    # 5
    {
        "filename": "RS_metro_parking_2025.xlsx",
        "insured": "Metro Parking Structures Inc.",
        "occupancy": "Real Estate Portfolio",
        "territory": "California",
        "tiv": 120_000_000, "locations": 8,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Declined", "year": 2025,
        "uw_notes": (
            "Metro Parking -- 8 multi-storey parking structures in LA and SF. "
            "TIV $120M. DECLINE.\n\n"
            "Parking structures in California present severe seismic exposure. "
            "6 of 8 structures are pre-1994 Northridge earthquake code, and 3 "
            "have NOT been retrofitted. EQ PML at 500yr return is $72M -- "
            "essentially 60% of TIV.\n\n"
            "Additionally, two structures have EV charging stations without "
            "adequate fire suppression -- lithium battery fire risk in enclosed "
            "parking is an emerging concern we're monitoring closely.\n\n"
            "Gallagher (Diane Murphy) is seeking $4.00/thousand but the EQ "
            "exposure alone requires $7.50+. Pre-Northridge unreinforced parking "
            "structures in CA are not insurable at market rates. If they complete "
            "seismic retrofits on the 3 deficient structures, we could revisit.\n\n"
            "Decline."
        ),
    },
    # 6
    {
        "filename": "RS_bayou_chemical_2024.xlsx",
        "insured": "Bayou Chemical Processing Corp.",
        "occupancy": "Chemical Processing",
        "territory": "Gulf Coast US",
        "tiv": 75_000_000, "locations": 2,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Declined", "year": 2024,
        "uw_notes": (
            "Bayou Chemical -- 2 chemical processing plants in LA (Baton Rouge, "
            "Lake Charles). TIV $75M. DECLINE.\n\n"
            "Chemical processing is inherently high-hazard and these facilities "
            "handle Class 1 flammable liquids. The Baton Rouge plant processes "
            "ethylene derivatives -- HPO (highly protected occupancy) standards "
            "are not met. No FM Global certification, no IRI inspection in 3+ "
            "years.\n\n"
            "Lake Charles is squarely in the wind corridor -- CAT load is "
            "punishing. Combined with the chemical hazard, the risk profile "
            "exceeds our appetite at any price point.\n\n"
            "Two claims: $450K process vessel fire (2023) and $220K hurricane "
            "Laura residual damage (still developing from 2020). 5-year loss "
            "ratio 56%.\n\n"
            "BMS Group (Marcus Wright) understands -- this needs a specialist "
            "chemical facility market, not a commercial property generalist. "
            "Decline."
        ),
    },
    # 7
    {
        "filename": "RS_sunshine_cannabis_2025.xlsx",
        "insured": "Sunshine Cannabis Cultivation Inc.",
        "occupancy": "Agricultural",
        "territory": "California",
        "tiv": 30_000_000, "locations": 4,
        "construction": "Frame (ISO Class 1)",
        "decision": "Declined", "year": 2025,
        "uw_notes": (
            "Sunshine Cannabis -- 4 cultivation and processing facilities in "
            "northern CA. TIV $30M. DECLINE.\n\n"
            "Cannabis operations remain outside our underwriting appetite as a "
            "matter of corporate policy. Regardless of state legality, federal "
            "classification creates reinsurance complications. Additionally:\n\n"
            "1. Frame construction, 2 of 4 unsprinklered\n"
            "2. High-intensity electrical systems for indoor growing -- fire "
            "risk is significant\n"
            "3. Remote locations in Humboldt/Mendocino -- wildfire zone and "
            "limited fire response\n"
            "4. Product liability and regulatory risk complicate the property "
            "coverage\n\n"
            "Aon (David Morris) knows our position on cannabis but submitted "
            "anyway hoping the policy had changed. It hasn't. Decline on "
            "underwriting appetite grounds."
        ),
    },
    # 8
    {
        "filename": "RS_eastside_plastics_2025.xlsx",
        "insured": "Eastside Plastics Manufacturing",
        "occupancy": "Heavy Manufacturing",
        "territory": "Midwest US",
        "tiv": 20_000_000, "locations": 1,
        "construction": "Joisted Masonry (ISO Class 3)",
        "decision": "Declined", "year": 2025,
        "uw_notes": (
            "Eastside Plastics -- single facility plastics moulding and extrusion "
            "in Cleveland, OH. TIV $20M. DECLINE.\n\n"
            "Plastics manufacturing is a challenging class due to combustible raw "
            "materials and high-temperature processes. This facility compounds the "
            "risk with: joisted masonry construction (1972), no compartmentation "
            "between production and warehouse, and inadequate sprinkler system "
            "(the existing wet-pipe system was designed for ordinary hazard but "
            "the occupancy is extra hazard).\n\n"
            "Three claims in 4 years: $180K extruder fire (2024), $95K raw material "
            "fire (2023), $60K electrical (2022). Loss ratio 105%.\n\n"
            "Willis (Robert Chen) has been told we need: (a) sprinkler upgrade to "
            "EH density, (b) fire walls between production and storage, and "
            "(c) 2 clean years before we'd reconsider. Decline for now."
        ),
    },
    # 9
    {
        "filename": "RS_island_resort_2024.xlsx",
        "insured": "Island Resort Collection Ltd.",
        "occupancy": "Hotel / Hospitality",
        "territory": "Caribbean",
        "tiv": 250_000_000, "locations": 5,
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "decision": "Declined", "year": 2024,
        "uw_notes": (
            "Island Resort Collection -- 5 luxury resort properties across "
            "Bahamas, Turks & Caicos, and BVI. TIV $250M. DECLINE.\n\n"
            "Caribbean hospitality portfolios are essentially CAT plays and "
            "our appetite for this class is effectively zero post-Ian. The "
            "250yr PML on this portfolio is $145M. Named storm deductibles "
            "help but don't change the fundamental exposure.\n\n"
            "The properties themselves are well-built (masonry non-combustible, "
            "hurricane-rated windows, generators) but location is the dominant "
            "factor. No loss history available -- this is a new-to-market risk "
            "that previously self-insured through a captive.\n\n"
            "Lockton (Emily Richards) was transparent about the challenge. She's "
            "trying to build a multi-carrier programme. Even at lead-follow terms, "
            "we can't allocate CAT capacity to Caribbean hospitality. The maths "
            "simply don't work for our book. Decline."
        ),
    },
    # 10
    {
        "filename": "RS_rapidfire_logistics_2025.xlsx",
        "insured": "Rapidfire Logistics Ltd.",
        "occupancy": "Warehouse / Distribution",
        "territory": "Midlands UK",
        "tiv": 48_000_000, "locations": 3,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Declined", "year": 2025,
        "uw_notes": (
            "Rapidfire Logistics -- 3 large distribution centres near Birmingham "
            "and Coventry. TIV GBP 38M (approx $48M). DECLINE.\n\n"
            "On paper this looks standard but the survey report raised serious "
            "concerns: (1) ESFR sprinkler system at the Coventry DC failed its "
            "last flow test and hasn't been remedied, (2) racking heights exceed "
            "sprinkler design -- goods stored above 12m in a system designed for "
            "10.5m max, (3) housekeeping noted as poor -- combustible packaging "
            "materials accumulating in aisles.\n\n"
            "The surveyor flagged all 3 issues in their report dated Jan 2025. "
            "We contacted the broker (Howden, Helen Taylor) who confirmed the "
            "insured is aware but hasn't committed to remediation timeline.\n\n"
            "A distribution centre with a deficient sprinkler system and storage "
            "exceeding design parameters is an unacceptable risk. If they rectify "
            "all 3 survey findings and provide evidence, we'll re-quote. Until "
            "then, decline."
        ),
    },
]

NTU_ACCOUNTS = [
    {
        "filename": "RS_greenfield_office_2025.xlsx",
        "insured": "Greenfield Office Park LLC",
        "occupancy": "Commercial Office",
        "territory": "Northeast US",
        "tiv": 60_000_000, "locations": 4,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "NTU", "year": 2025,
        "uw_notes": (
            "Greenfield Office Park -- 4 suburban office buildings in NJ. TIV $60M. "
            "NTU -- broker placed elsewhere.\n\n"
            "Clean risk, fire resistive, fully sprinklered, no claims. We quoted "
            "$2.60/thousand ($156K) against technical of $2.80. Aon (Lisa Chang) "
            "came back and said AIG offered $2.20/thousand.\n\n"
            "We're not matching $2.20 -- that's 21% below our technical rate and "
            "below what we believe is adequate for the class. AIG appears to be "
            "buying market share in suburban office, which is their prerogative, "
            "but we won't participate in a race to the bottom.\n\n"
            "If AIG's pricing doesn't hold at renewal and this comes back, we'd "
            "re-quote at our original $2.60 or wherever technical sits at that "
            "point. No hard feelings -- Lisa knows our position."
        ),
    },
    {
        "filename": "RS_pioneer_manufacturing_2025.xlsx",
        "insured": "Pioneer Manufacturing Corp.",
        "occupancy": "Light Manufacturing",
        "territory": "Midwest US",
        "tiv": 35_000_000, "locations": 2,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "NTU", "year": 2025,
        "uw_notes": (
            "Pioneer Manufacturing -- 2 facilities in WI. TIV $35M. NTU.\n\n"
            "Good risk that we wanted to write. Light manufacturing (injection "
            "moulding), non-combustible, sprinklered, 1 claim in 5 years ($40K). "
            "Quoted at $4.40/thousand ($154K). Technical was $4.60.\n\n"
            "Marsh (Tom Bradley) ran a full marketing exercise and Liberty Mutual "
            "came in at $3.80/thousand. That's 17% below our quote and 18% below "
            "technical. We offered to sharpen to $4.20 but Liberty's $3.80 was "
            "too aggressive a gap to bridge.\n\n"
            "Disappointing because this is exactly the risk profile we target. "
            "But we can't match pricing that we believe is inadequate. Tom said "
            "he'll keep us on the panel for renewal."
        ),
    },
    {
        "filename": "RS_thames_valley_tech_2024.xlsx",
        "insured": "Thames Valley Technology Park Ltd.",
        "occupancy": "Commercial Office",
        "territory": "South East England",
        "tiv": 50_000_000, "locations": 3,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "NTU", "year": 2024,
        "uw_notes": (
            "Thames Valley Tech Park -- 3 tech office campuses near Reading, UK. "
            "TIV GBP 40M (approx $50M). NTU.\n\n"
            "Modern, fire resistive, fully sprinklered, no loss history. We quoted "
            "GBP 2.40/thousand. Ed Broking (Natalie Ford) told us Zurich came in "
            "at GBP 1.90/thousand on a follow basis.\n\n"
            "GBP 1.90 is significantly below where we see technical for this class, "
            "even acknowledging it's a clean risk. We suspect Zurich is pricing for "
            "the broader tech sector relationship rather than the individual risk.\n\n"
            "Not matching. If Zurich doesn't renew at that level, we'd be happy "
            "to re-engage. NTU."
        ),
    },
    {
        "filename": "RS_heartland_foods_2025.xlsx",
        "insured": "Heartland Foods Processing Inc.",
        "occupancy": "Food Processing",
        "territory": "Midwest US",
        "tiv": 30_000_000, "locations": 2,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "NTU", "year": 2025,
        "uw_notes": (
            "Heartland Foods -- 2 meat processing facilities in NE and KS. TIV "
            "$30M. NTU -- placed with incumbent.\n\n"
            "Food processing is a class we're building expertise in. Quoted at "
            "$5.80/thousand ($174K) which was competitive with technical at $6.00. "
            "However, the insured chose to stay with their incumbent (Travelers) "
            "at $5.40/thousand with multi-year deal.\n\n"
            "Multi-year commitments aren't something we typically offer on "
            "property, and matching a 3-year rate lock at $5.40 isn't justified "
            "by the risk profile. The LTR (long-term relationship) between "
            "Heartland and Travelers goes back 8 years.\n\n"
            "Not surprised. Willis (Emily Watson) gave us a fair shot but the "
            "incumbent relationship was too strong. We'll see this again if "
            "Travelers adjusts pricing at the 3-year renewal."
        ),
    },
    {
        "filename": "RS_summit_retail_2024.xlsx",
        "insured": "Summit Retail Holdings LLC",
        "occupancy": "Retail",
        "territory": "Mid-Atlantic US",
        "tiv": 45_000_000, "locations": 8,
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "decision": "NTU", "year": 2024,
        "uw_notes": (
            "Summit Retail Holdings -- 8 shopping centre locations in PA and MD. "
            "TIV $45M. NTU -- submission withdrawn.\n\n"
            "We were in the process of quoting when Howden (Michael Cross) "
            "informed us the insured decided to restructure their programme as "
            "a captive arrangement with excess only placement. Our quote for "
            "the primary layer was $4.20/thousand.\n\n"
            "The captive restructuring changes the risk profile entirely -- "
            "we'd need to re-underwrite for excess attachment. Michael said "
            "they may come back for the excess layer but no timeline.\n\n"
            "File closed as NTU. No further action required unless they "
            "re-submit."
        ),
    },
    {
        "filename": "RS_bluestone_warehouse_2025.xlsx",
        "insured": "Bluestone Warehousing Group",
        "occupancy": "Warehouse / Distribution",
        "territory": "North West England",
        "tiv": 35_000_000, "locations": 4,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "NTU", "year": 2025,
        "uw_notes": (
            "Bluestone Warehousing -- 4 distribution centres in Manchester and "
            "Liverpool area. TIV GBP 28M (approx $35M). NTU.\n\n"
            "Standard warehousing risk, non-combustible, sprinklered. We quoted "
            "GBP 3.20/thousand (GBP 89.6K). Broker came back saying Aviva offered "
            "GBP 2.60/thousand on a 2-year LTA.\n\n"
            "Can't compete with Aviva on a 2-year LTA at that rate. GBP 2.60 is "
            "below our technical for warehousing in the North West, even with clean "
            "experience.\n\n"
            "Miller Insurance (Charlotte Webb) says she'll bring this back if Aviva "
            "corrects at renewal. Not unexpected -- Aviva is aggressive on UK "
            "warehouse currently."
        ),
    },
    {
        "filename": "RS_cardinal_health_2024.xlsx",
        "insured": "Cardinal Healthcare Properties LP",
        "occupancy": "Healthcare / Hospital",
        "territory": "Southeast US",
        "tiv": 95_000_000, "locations": 5,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "NTU", "year": 2024,
        "uw_notes": (
            "Cardinal Healthcare -- 5 hospital facilities in GA and FL. TIV $95M. "
            "NTU -- insured moved to a group captive programme.\n\n"
            "We quoted at $3.40/thousand ($323K). Competitive quote, within "
            "technical. The insured was satisfied with our terms but their parent "
            "company made a strategic decision to move all healthcare properties "
            "into a shared captive.\n\n"
            "This was not a pricing issue -- BMS Group (Samantha Lee) confirmed "
            "our quote was the most competitive of the 4 markets approached. "
            "Corporate decision overrode the local insurance buyer.\n\n"
            "Good risk that we'd have been happy to write. If the captive "
            "arrangement doesn't work out, we're first call. NTU."
        ),
    },
    {
        "filename": "RS_oak_ridge_2025.xlsx",
        "insured": "Oak Ridge Business Centers LLC",
        "occupancy": "Commercial Office",
        "territory": "Southeast US",
        "tiv": 25_000_000, "locations": 3,
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "decision": "NTU", "year": 2025,
        "uw_notes": (
            "Oak Ridge Business Centers -- 3 suburban office parks in Atlanta metro. "
            "TIV $25M. NTU.\n\n"
            "Straightforward suburban office risk. Masonry non-combustible, sprinklered, "
            "no claims. We quoted $2.80/thousand ($70K). Gallagher (Kevin O'Brien) "
            "informed us Cincinnati Financial came in at $2.30/thousand on a package "
            "with GL and auto.\n\n"
            "We don't write casualty so we can't compete on a multi-line package "
            "basis. The property-only rate of $2.30 is achievable only because "
            "Cincinnati is cross-subsidizing from the casualty lines.\n\n"
            "This is a recurring challenge with smaller accounts -- package "
            "writers will always have an advantage when the insured wants "
            "convenience. NTU. No further action."
        ),
    },
]

REFERRED_ACCOUNTS = [
    {
        "filename": "RS_grandview_hotel_2025.xlsx",
        "insured": "Grandview Hotel Corporation",
        "occupancy": "Hotel / Hospitality",
        "territory": "Florida",
        "tiv": 140_000_000, "locations": 5,
        "construction": "Masonry Non-Combustible (ISO Class 5)",
        "decision": "Referred", "year": 2025,
        "uw_notes": (
            "Grandview Hotel -- 5 resort hotels, 3 in FL (Orlando, Tampa, Destin), "
            "1 in Hilton Head SC, 1 in Myrtle Beach SC. TIV $140M. REFERRED to "
            "Simon Thornton -- above Level 2 authority.\n\n"
            "The Orlando and Tampa properties are inland and acceptable. But the "
            "Destin, Hilton Head, and Myrtle Beach locations bring material wind "
            "exposure. Aggregate PML at 250yr is $52M.\n\n"
            "Technical rate $6.20/thousand, broker seeking $5.80. Deviation is "
            "6.5% which would normally be within my authority, but the TIV exceeds "
            "my $100M threshold.\n\n"
            "My recommendation to Simon: bind at $5.80 with named storm deductible "
            "of 3% per coastal location and 2% aggregate. The inland locations "
            "anchor the portfolio and the blended rate is adequate. Need fac for "
            "the wind excess layer.\n\n"
            "Aon (Paul Henderson) is pushing for quick turnaround. Referred for "
            "Simon's decision."
        ),
    },
    {
        "filename": "RS_westminster_mixed_2025.xlsx",
        "insured": "Westminster Property Holdings Plc",
        "occupancy": "Mixed-Use Commercial",
        "territory": "Central London",
        "tiv": 300_000_000, "locations": 12,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "Referred", "year": 2025,
        "uw_notes": (
            "Westminster Property Holdings -- 12 premium mixed-use properties in "
            "central London. TIV GBP 240M (approx $300M). REFERRED -- well above "
            "any individual authority level.\n\n"
            "This is a marquee account. Grade A office, luxury retail, and "
            "residential across Mayfair, St James's, Knightsbridge. All fire "
            "resistive, fully sprinklered, 24/7 security. No loss history on "
            "record.\n\n"
            "The technical rate GBP 2.20/thousand generates a premium of GBP 528K "
            "($660K). Broker targeting GBP 2.00/thousand. The size alone requires "
            "board-level approval and a significant fac placement.\n\n"
            "We'd be taking 15% line at most -- GBP 79.2K our share. The attraction "
            "is prestige and relationship with Guy Carpenter (Sophie Jenkins) who "
            "controls a substantial London commercial portfolio.\n\n"
            "My recommendation: participate at 15% behind a strong lead (likely "
            "Hiscox or Beazley). Need Simon and the Head of Property to approve. "
            "Referred."
        ),
    },
    {
        "filename": "RS_tristate_industrial_2025.xlsx",
        "insured": "Tristate Industrial Complex LLC",
        "occupancy": "Heavy Manufacturing",
        "territory": "Northeast US",
        "tiv": 110_000_000, "locations": 4,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Referred", "year": 2025,
        "uw_notes": (
            "Tristate Industrial -- 4 manufacturing facilities in NJ and PA. "
            "TIV $110M. REFERRED -- above Level 2 authority on TIV.\n\n"
            "Heavy manufacturing (steel fabrication, welding, coating) is a "
            "class that requires careful selection. These facilities are well-"
            "managed: all sprinklered, proper hot work programmes, NFPA-"
            "compliant electrical. FM Global certified.\n\n"
            "One claim in 5 years: $150K welding fire (2024), contained by "
            "sprinklers. Loss ratio 9%. Technical rate $5.40/thousand, broker "
            "seeking $5.00 (7.4% deviation).\n\n"
            "My view: the risk quality justifies the deviation. FM certification "
            "is a strong indicator of good practice. Comparable to Allegheny "
            "Steel Works (2023) bound at $5.20.\n\n"
            "Recommending to Simon: bind at $5.00/thousand ($550K) with hot work "
            "warranty and annual FM inspection certification. Need fac for layers "
            "above $30M retention."
        ),
    },
    {
        "filename": "RS_pacific_ports_2025.xlsx",
        "insured": "Pacific Ports & Terminals Inc.",
        "occupancy": "Warehouse / Distribution",
        "territory": "California",
        "tiv": 200_000_000, "locations": 3,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Referred", "year": 2025,
        "uw_notes": (
            "Pacific Ports -- 3 marine terminal and warehouse facilities in Long "
            "Beach and Oakland. TIV $200M. REFERRED -- TIV significantly above "
            "authority and EQ exposure requires senior review.\n\n"
            "Marine terminal is a specialty class. The warehousing component is "
            "straightforward but the wharf and dock structures present unique "
            "exposures. EQ PML at 250yr is $65M due to liquefaction risk at "
            "the Long Beach facilities.\n\n"
            "No loss history (new to market). Technical rate $5.80/thousand, "
            "broker seeking $5.00 (13.8% deviation). That's a significant gap.\n\n"
            "My recommendation: if we participate, it should be at technical "
            "rate ($5.80) with EQ deductible of 5% and our line limited to 10%. "
            "This caps our exposure at $20M net of which EQ maximum is $5.5M "
            "after deductible.\n\n"
            "McGill (Derek Stone) is building a multi-carrier tower. Referred "
            "to Simon for approval of terms and line size."
        ),
    },
    {
        "filename": "RS_nexus_data_2025.xlsx",
        "insured": "Nexus Data Systems Inc.",
        "occupancy": "Data Centre",
        "territory": "Mid-Atlantic US",
        "tiv": 250_000_000, "locations": 3,
        "construction": "Fire Resistive (ISO Class 6)",
        "decision": "Referred", "year": 2025,
        "uw_notes": (
            "Nexus Data Systems -- 3 hyperscale data centres in northern VA. "
            "TIV $250M. REFERRED -- TIV well above authority.\n\n"
            "Data centres are our target growth class and this is a top-tier "
            "facility operator. Tier IV design, concurrently maintainable, "
            "N+2 power redundancy, clean agent suppression, seismically "
            "engineered. The facilities are among the best-protected in our "
            "portfolio.\n\n"
            "No loss history. Technical rate $3.40/thousand ($850K technical "
            "premium). Broker (Gallagher, Rachel Foster) seeking $3.00. "
            "Deviation is 11.8% which is significant, but data centre "
            "competition is fierce.\n\n"
            "Comparable: Cascade Tech (2026) bound at $3.50, Richmond DC "
            "(2025) at $3.40. The $3.00 ask is below recent comparables.\n\n"
            "My recommendation: offer at $3.20/thousand with 20% line ($800K "
            "total, $160K our share). This maintains rate integrity while "
            "staying competitive. Rachel has indicated $3.20 could work. "
            "Referring to Simon for the TIV authority approval."
        ),
    },
]

DETERIORATING_RENEWALS = [
    {
        "filename": "RS_eastcoast_food_2026.xlsx",
        "insured": "Eastcoast Food Services Inc.",
        "occupancy": "Food Processing",
        "territory": "Northeast US",
        "tiv": 38_000_000, "locations": 2,
        "construction": "Joisted Masonry (ISO Class 3)",
        "decision": "Bound",
        "year": 2026,
        "uw_notes": (
            "Eastcoast Food Services renewal -- third year on the book and the "
            "loss experience is trending badly. Year 1 was clean. Year 2 had a "
            "$95K grease duct fire (kitchen exhaust system) and a $45K slip/trip "
            "water damage claim. Year 3 (current) has a $280K walk-in freezer "
            "failure with full inventory loss and a $60K fire in the packaging "
            "area.\n\n"
            "Cumulative losses: $480K over 3 years. Premiums collected: $390K. "
            "Loss ratio: 123%. This is unsustainable.\n\n"
            "We're renewing but at a significant rate increase: expiring $5.40/"
            "thousand ($205.2K), renewing at $7.00/thousand ($266K) -- a 30% "
            "increase. The alternative was non-renewal but the broker (Marsh, "
            "James Whitfield) made a convincing case that the freezer failure "
            "was a one-off mechanical event and the insured has since installed "
            "redundant cooling.\n\n"
            "CONDITIONS for renewal: (1) $100K deductible (up from $25K), (2) "
            "quarterly kitchen exhaust cleaning certification, (3) redundant "
            "freezer alarm system. If there's ANOTHER large loss in year 4, "
            "we non-renew without discussion. James understands this.\n\n"
            "Referred to Simon for the rate adequacy question. Simon approved "
            "with the conditions noted."
        ),
    },
    {
        "filename": "RS_central_warehouse_2026.xlsx",
        "insured": "Central Warehouse Solutions Ltd.",
        "occupancy": "Warehouse / Distribution",
        "territory": "Midlands UK",
        "tiv": 52_000_000, "locations": 3,
        "construction": "Non-Combustible (ISO Class 4)",
        "decision": "Bound",
        "year": 2026,
        "uw_notes": (
            "Central Warehouse Solutions renewal -- 2nd year. TIV GBP 41M (approx "
            "$52M). 3 distribution centres in the West Midlands.\n\n"
            "Year 1 loss experience was acceptable (GBP 35K forklift damage, GBP "
            "22K roof leak) but Q4 2025 brought a GBP 420K fire in the Wolverhampton "
            "DC -- electrical fault in the goods-in area that spread to racked "
            "stock before sprinklers contained it. Investigation found the sprinkler "
            "heads in that zone were obstructed by non-compliant racking installed "
            "by a contractor without notifying us.\n\n"
            "This is a compliance failure by the insured. The fire SHOULD have been "
            "a GBP 50K event if sprinklers had operated as designed. Instead it was "
            "GBP 420K. Still open with GBP 80K reserved.\n\n"
            "Renewing at GBP 4.80/thousand (up from GBP 3.60 -- a 33% increase). "
            "CONDITIONS: (1) immediate racking compliance audit across all 3 DCs, "
            "(2) sprinkler system re-certification, (3) written confirmation that "
            "no racking modifications will be made without insurer notification.\n\n"
            "Howden (Helen Taylor) has been straightforward about the contractor "
            "issue. The insured is cooperating fully. But trust has been damaged "
            "and the rate increase reflects both the loss and the compliance breach. "
            "If the audit reveals further sprinkler obstructions, we'll need to "
            "reconsider our position entirely."
        ),
    },
]


# ---------------------------------------------------------------------------
# Helper: build a complete record from a skeleton
# ---------------------------------------------------------------------------
def complete_record(skel, idx):
    """Fill in missing fields for non-hero accounts."""
    random.seed(42 + idx)
    rec = dict(skel)

    year = rec.get("year", 2025)
    month = random.choice([1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12])
    rec.setdefault("eff_date", datetime.date(year, month, 1))
    rec.setdefault("exp_date", datetime.date(year + 1, month, 1))

    tiv = rec["tiv"]
    territory = rec["territory"]
    occ = rec["occupancy"]
    decision = rec["decision"]

    # pick broker
    firm = random.choice(list(BROKERS_FIRMS.keys()))
    contact = random.choice(BROKERS_FIRMS[firm])
    rec.setdefault("broker_firm", firm)
    rec.setdefault("broker_contact", contact)

    rec.setdefault("underwriter", random.choice(UNDERWRITERS))

    # rating factors
    base_rate = round(random.uniform(2.0, 8.5), 2)
    terr_factor = TERRITORIES.get(territory, 1.00)
    occ_factor = round(random.uniform(0.85, 1.20), 2)
    const_factor = round(random.uniform(0.80, 1.10), 2)
    prot_factor = round(random.uniform(0.80, 1.15), 2)
    exp_mod = round(random.uniform(0.85, 1.20), 2)
    sched_mod = round(random.uniform(0.95, 1.05), 2)
    uw_judg = round(random.uniform(0.95, 1.10), 2)
    pkg_credit = round(random.uniform(0.95, 1.00), 2)

    tech_rate = round(base_rate * terr_factor * occ_factor * const_factor *
                      prot_factor * exp_mod * sched_mod * uw_judg * pkg_credit /
                      (terr_factor * occ_factor * const_factor * prot_factor *
                       exp_mod * sched_mod * uw_judg * pkg_credit) * base_rate /
                      base_rate * random.uniform(0.95, 1.05) * base_rate /
                      base_rate, 2)
    # Simplify -- just use a sensible technical rate derived from base
    tech_rate = round(base_rate * random.uniform(0.90, 1.10), 2)

    if decision == "Bound":
        quoted_rate = round(tech_rate * random.uniform(0.88, 1.02), 2)
        line_pct = random.choice([0.15, 0.20, 0.25, 0.50, 1.00])
    elif decision == "Declined":
        # broker asking way below technical
        quoted_rate = round(tech_rate * random.uniform(0.55, 0.80), 2)
        line_pct = 0.0
    elif decision == "NTU":
        quoted_rate = round(tech_rate * random.uniform(0.90, 1.00), 2)
        line_pct = 0.0
    else:  # Referred
        quoted_rate = round(tech_rate * random.uniform(0.85, 0.98), 2)
        line_pct = random.choice([0.15, 0.20, 0.25])

    tech_premium = round(tiv / 1000 * tech_rate)
    our_premium = round(tiv / 1000 * quoted_rate * line_pct)

    rec.setdefault("coverage", "All-Risk Property incl. BI")
    rec.setdefault("our_line_pct", line_pct)
    rec.setdefault("our_premium", our_premium)
    rec.setdefault("technical_premium", tech_premium)
    rec.setdefault("decision_date", rec["eff_date"] - datetime.timedelta(days=random.randint(14, 60)))
    rec.setdefault("base_rate", base_rate)
    rec.setdefault("territory_factor", terr_factor)
    rec.setdefault("occupancy_factor", occ_factor)
    rec.setdefault("construction_factor", const_factor)
    rec.setdefault("protection_factor", prot_factor)
    rec.setdefault("experience_mod", exp_mod)
    rec.setdefault("schedule_mod", sched_mod)
    rec.setdefault("uw_judgment", uw_judg)
    rec.setdefault("package_credit", pkg_credit)
    rec.setdefault("technical_rate", tech_rate)
    rec.setdefault("quoted_rate", quoted_rate)
    rec.setdefault("cat_wind", round(random.uniform(0.05, 3.00), 2) if "FL" in territory or "Gulf" in territory or "Caribbean" in territory else round(random.uniform(0.05, 0.50), 2))
    rec.setdefault("cat_eq", round(random.uniform(0.02, 0.40), 2) if "California" in territory else round(random.uniform(0.02, 0.15), 2))
    rec.setdefault("cat_flood", round(random.uniform(0.03, 0.30), 2))
    rec.setdefault("expense_ratio", round(random.uniform(0.28, 0.35), 2))
    rec.setdefault("profit_margin", 0.05)

    # loss years
    if "loss_years" not in rec:
        loss_years = []
        for yr in range(year - 5, year):
            n_claims = random.choices([0, 1, 2, 3], weights=[50, 30, 15, 5])[0]
            total_incurred = sum(random.randint(10_000, 400_000) for _ in range(n_claims)) if n_claims else 0
            approx_premium = tiv / 1000 * tech_rate * 0.9
            lr = round(total_incurred / approx_premium, 2) if approx_premium > 0 else 0.0
            loss_years.append((yr, n_claims, total_incurred, lr))
        rec["loss_years"] = loss_years

    rec.setdefault("large_losses", "No individual loss exceeds reporting threshold.")
    rec.setdefault("loss_dev_notes", "All claims fully developed. No adverse development.")

    return rec


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
def build_workbook(rec):
    wb = Workbook()

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    ws.cell(row=1, column=1, value="RATING DATA SHEET").font = TITLE_FONT
    ws.merge_cells("A1:B1")

    r = 3
    fields = [
        ("Insured Name", rec["insured"]),
        ("Broker", f"{rec['broker_firm']} ({rec['broker_contact']})"),
        ("Effective Date", rec["eff_date"]),
        ("Expiry Date", rec["exp_date"]),
        ("Coverage Type", rec["coverage"]),
        ("Total Insured Value (TIV)", rec["tiv"]),
        ("Number of Locations", rec["locations"]),
        ("Our Line %", rec["our_line_pct"]),
        ("Our Premium", rec["our_premium"]),
        ("Technical Premium", rec["technical_premium"]),
        ("Decision", rec["decision"]),
        ("Decision Date", rec["decision_date"]),
        ("Underwriter", rec["underwriter"]),
    ]
    for label, val in fields:
        write_label_value(ws, r, label, val)
        if "Premium" in label or "Value" in label:
            ws.cell(row=r, column=2).number_format = '#,##0'
        if "Line %" in label:
            ws.cell(row=r, column=2).number_format = '0.00%'
        r += 1

    # ---- Sheet 2: Rating ----
    ws2 = wb.create_sheet("Rating")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    ws2.cell(row=1, column=1, value="RATING ANALYSIS").font = TITLE_FONT
    ws2.merge_cells("A1:B1")

    r = 3
    rating_fields = [
        ("Base Rate (per $1,000)", rec["base_rate"], '0.00'),
        ("Territory Factor", rec["territory_factor"], '0.00'),
        ("Occupancy Factor", rec["occupancy_factor"], '0.00'),
        ("Construction Factor", rec["construction_factor"], '0.00'),
        ("Protection Factor", rec["protection_factor"], '0.00'),
        ("Experience Modification", rec["experience_mod"], '0.00'),
        ("Schedule Modification", rec["schedule_mod"], '0.00'),
        ("UW Judgment Factor", rec["uw_judgment"], '0.00'),
        ("Package Credit", rec["package_credit"], '0.00'),
    ]
    for label, val, fmt in rating_fields:
        write_label_value(ws2, r, label, val, fmt=fmt)
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="PREMIUM SUMMARY").font = LABEL_FONT
    r += 1
    tech_rate = rec["technical_rate"]
    quoted_rate = rec["quoted_rate"]
    tiv = rec["tiv"]
    tech_prem = rec["technical_premium"]
    quoted_prem = round(tiv / 1000 * quoted_rate)
    deviation = round((quoted_rate - tech_rate) / tech_rate * 100, 1) if tech_rate else 0

    prem_fields = [
        ("Technical Rate (per $1,000)", tech_rate, '0.00'),
        ("Technical Premium", tech_prem, '#,##0'),
        ("Quoted Rate (per $1,000)", quoted_rate, '0.00'),
        ("Quoted Premium", quoted_prem, '#,##0'),
        ("Deviation %", deviation, '0.0'),
    ]
    for label, val, fmt in prem_fields:
        write_label_value(ws2, r, label, val, fmt=fmt)
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="CAT LOADS (per $1,000)").font = LABEL_FONT
    r += 1
    cat_fields = [
        ("Wind", rec["cat_wind"], '0.00'),
        ("Earthquake", rec["cat_eq"], '0.00'),
        ("Flood", rec["cat_flood"], '0.00'),
    ]
    for label, val, fmt in cat_fields:
        write_label_value(ws2, r, label, val, fmt=fmt)
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="EXPENSE & MARGIN").font = LABEL_FONT
    r += 1
    write_label_value(ws2, r, "Expense Ratio", rec["expense_ratio"], fmt='0.00%')
    r += 1
    write_label_value(ws2, r, "Profit Margin", rec["profit_margin"], fmt='0.00%')

    # ---- Sheet 3: Loss Analysis ----
    ws3 = wb.create_sheet("Loss Analysis")
    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 15

    ws3.cell(row=1, column=1, value="LOSS ANALYSIS").font = TITLE_FONT
    ws3.merge_cells("A1:D1")

    r = 3
    ws3.cell(row=r, column=1, value="5-Year Loss Summary").font = LABEL_FONT
    r += 1
    headers = ["Policy Year", "# Claims", "Total Incurred", "Loss Ratio"]
    for c, h in enumerate(headers, 1):
        ws3.cell(row=r, column=c, value=h)
    style_header_row(ws3, r, 4)
    r += 1

    for yr, nc, ti, lr in rec["loss_years"]:
        ws3.cell(row=r, column=1, value=yr)
        style_data_cell(ws3, r, 1)
        ws3.cell(row=r, column=2, value=nc)
        style_data_cell(ws3, r, 2)
        ws3.cell(row=r, column=3, value=ti)
        style_data_cell(ws3, r, 3, fmt='#,##0')
        ws3.cell(row=r, column=4, value=lr)
        style_data_cell(ws3, r, 4, fmt='0.00')
        r += 1

    r += 2
    ws3.cell(row=r, column=1, value="Large Loss Detail").font = LABEL_FONT
    r += 1
    ws3.cell(row=r, column=1, value=rec["large_losses"]).font = VALUE_FONT
    ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)

    r += 4
    ws3.cell(row=r, column=1, value="Loss Development Notes").font = LABEL_FONT
    r += 1
    ws3.cell(row=r, column=1, value=rec["loss_dev_notes"]).font = VALUE_FONT
    ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=4)

    # ---- Sheet 4: UW Notes ----
    ws4 = wb.create_sheet("UW Notes")
    ws4.column_dimensions["A"].width = 120

    ws4.cell(row=1, column=1, value="UNDERWRITER COMMENTARY").font = TITLE_FONT

    note_cell = ws4.cell(row=3, column=1, value=rec["uw_notes"])
    note_cell.font = NOTE_FONT
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Merge a generous area so the note is readable
    ws4.merge_cells("A3:A50")

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    all_records = []

    # 10 hero accounts
    for h in HERO_ACCOUNTS:
        all_records.append(h)

    # 15 historical bound
    for i, skel in enumerate(HISTORICAL_BOUND):
        all_records.append(complete_record(skel, i))

    # 10 historical declined
    for i, skel in enumerate(HISTORICAL_DECLINED):
        all_records.append(complete_record(skel, 100 + i))

    # 8 NTU
    for i, skel in enumerate(NTU_ACCOUNTS):
        all_records.append(complete_record(skel, 200 + i))

    # 5 referred
    for i, skel in enumerate(REFERRED_ACCOUNTS):
        all_records.append(complete_record(skel, 300 + i))

    # 2 deteriorating renewals
    for i, skel in enumerate(DETERIORATING_RENEWALS):
        all_records.append(complete_record(skel, 400 + i))

    print(f"Generating {len(all_records)} rating data sheets...")

    for rec in all_records:
        wb = build_workbook(rec)
        path = os.path.join(OUTPUT_DIR, rec["filename"])
        wb.save(path)
        print(f"  Created: {rec['filename']}  [{rec['decision']}]")

    print(f"\nDone. {len(all_records)} workbooks saved to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
