"""
Generate 12 realistic Schedule of Values (SOV) Excel files
for Tokio Marine HCC commercial property underwriting POC.
"""

import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

random.seed(42)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "02_sov_schedules")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Standard column headers
# ---------------------------------------------------------------------------
STANDARD_HEADERS = [
    "Location Number", "Location Name", "Street Address", "City",
    "State/Country", "ZIP/Postal Code", "Building Description",
    "Occupancy Type", "Construction Type", "Year Built",
    "Number of Stories", "Square Footage", "Sprinkler System (Y/N)",
    "Sprinkler Type", "Fire Alarm (Central/Local/None)",
    "Security System (Y/N)", "Roof Type", "Roof Age",
    "Building Value", "Contents Value", "Business Interruption Value",
    "Total Insured Value", "Flood Zone (A/B/C/X)", "Wind Tier (1/2/3/4)",
    "Earthquake Zone (1-5)", "Distance to Coast (miles)",
    "Distance to Fire Station (miles)",
]

# Acme Industrial: missing sprinkler info & flood zone columns
ACME_HEADERS = [
    "Location Number", "Location Name", "Street Address", "City",
    "State/Country", "ZIP/Postal Code", "Building Description",
    "Occupancy Type", "Construction Type", "Year Built",
    "Number of Stories", "Square Footage",
    # -- sprinkler columns OMITTED --
    "Fire Alarm (Central/Local/None)",
    "Security System (Y/N)", "Roof Type", "Roof Age",
    "Building Value", "Contents Value", "Business Interruption Value",
    "Total Insured Value",
    # -- Flood Zone OMITTED --
    "Wind Tier (1/2/3/4)",
    "Earthquake Zone (1-5)", "Distance to Coast (miles)",
    "Distance to Fire Station (miles)",
]

# Midwest Ag: alternate header names
MIDWEST_HEADERS = [
    "Location Number", "Location Name", "Street Address", "City",
    "State/Country", "ZIP/Postal Code", "Building Description",
    "Occupancy Type", "Construction Type", "Year Built",
    "Number of Stories", "Square Footage", "Fire Protection",
    "Protection Type", "Alarm System",
    "Security System (Y/N)", "Roof Type", "Roof Age",
    "Building Cost", "Contents Cost", "BI Value",
    "Replacement Cost", "Flood Zone (A/B/C/X)", "Wind Tier (1/2/3/4)",
    "Earthquake Zone (1-5)", "Distance to Coast (miles)",
    "Distance to Fire Station (miles)",
]

# ---------------------------------------------------------------------------
# Helper: currency columns indices (0-based) for standard headers
# ---------------------------------------------------------------------------
CURRENCY_COLS_STD = [18, 19, 20, 21]   # Building, Contents, BI, TIV
CURRENCY_COLS_ACME = [16, 17, 18, 19]
CURRENCY_COLS_MIDWEST = [18, 19, 20, 21]


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------
def _rand_street():
    nums = random.randint(100, 9999)
    names = ["Main St", "Industrial Blvd", "Commerce Dr", "Oak Ave",
             "Maple Rd", "Park Ave", "Center St", "Market St",
             "Highway 1", "Route 9", "Elm St", "Pine Dr",
             "Broadway", "Lakeview Rd", "Airport Blvd", "Harbor Dr",
             "Ocean Ave", "River Rd", "Sunset Blvd", "First Ave",
             "Second St", "Third Ave", "Washington Blvd", "Lincoln Way",
             "Jefferson Dr", "Technology Pkwy", "Innovation Dr",
             "Enterprise Way", "Corporate Dr", "Logistics Ln",
             "Distribution Way", "Factory Rd", "Mill St",
             "Warehouse Row", "Terminal Blvd"]
    return f"{nums} {random.choice(names)}"


def _zip_for_state(state):
    zips = {
        "GA": ("30", 3), "FL": ("32", 3), "SC": ("29", 3),
        "NC": ("27", 3), "TN": ("37", 3), "AL": ("35", 3),
        "OH": ("44", 3), "PA": ("15", 3), "WA": ("98", 3),
        "CO": ("80", 3), "UT": ("84", 3), "AZ": ("85", 3),
        "NV": ("89", 3), "TX": ("77", 3), "MN": ("55", 3),
        "IA": ("50", 3), "WI": ("53", 3), "ND": ("58", 3),
        "MA": ("02", 3), "NY": ("10", 3), "NJ": ("07", 3),
        "CT": ("06", 3), "CA": ("90", 3),
    }
    prefix, digits = zips.get(state, ("00", 3))
    return prefix + "".join([str(random.randint(0, 9)) for _ in range(digits)])


def _zip_caribbean(country):
    return {
        "Bahamas": "BS-NP",
        "Turks & Caicos": "TKCA 1ZZ",
        "St Kitts": "KN-01",
        "Barbados": "BB15028",
    }.get(country, "00000")


def _distribute_tiv(total, n, min_pct=0.4, max_pct=1.6):
    """Distribute total across n locations with some variance."""
    avg = total / n
    raw = [avg * random.uniform(min_pct, max_pct) for _ in range(n)]
    factor = total / sum(raw)
    return [round(r * factor, -3) for r in raw]  # round to nearest $1000


def _split_tiv(tiv):
    """Split TIV into building, contents, BI."""
    bld_pct = random.uniform(0.50, 0.65)
    cnt_pct = random.uniform(0.15, 0.30)
    bi_pct = 1.0 - bld_pct - cnt_pct
    bld = round(tiv * bld_pct, -3)
    cnt = round(tiv * cnt_pct, -3)
    bi = tiv - bld - cnt
    return bld, cnt, bi


def _style_workbook(ws, headers, currency_cols, data_row_count):
    """Apply formatting: header style, column widths, currency format, totals."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="002060", end_color="002060",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Header row (row 1)
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row in range(2, 2 + data_row_count + 1):  # +1 for totals
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if (col - 1) in currency_cols:
                cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="center")

    # Totals row bold
    totals_row = 2 + data_row_count
    for col in range(1, len(headers) + 1):
        ws.cell(row=totals_row, column=col).font = Font(bold=True, size=11)

    # Column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in range(2, 2 + data_row_count + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # Freeze top row
    ws.freeze_panes = "A2"


def _totals_row(headers, rows, currency_cols):
    """Build a totals row summing currency columns."""
    totals = [""] * len(headers)
    totals[0] = ""
    totals[1] = "TOTALS"
    for ci in currency_cols:
        totals[ci] = sum(r[ci] for r in rows)
    # Sum square footage if present
    sqft_idx = None
    for i, h in enumerate(headers):
        if "square" in h.lower() or "sq" in h.lower():
            sqft_idx = i
            break
    if sqft_idx is not None:
        totals[sqft_idx] = sum(r[sqft_idx] for r in rows if isinstance(r[sqft_idx], (int, float)))
    return totals


def _write_workbook(filename, headers, rows, currency_cols, extra_sheets=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule of Values"

    # Write headers
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    # Write data
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    # Totals
    totals = _totals_row(headers, rows, currency_cols)
    tr = len(rows) + 2
    for ci, val in enumerate(totals, 1):
        ws.cell(row=tr, column=ci, value=val)

    _style_workbook(ws, headers, currency_cols, len(rows))

    # Extra sheets
    if extra_sheets:
        for sheet_name, sheet_data in extra_sheets.items():
            ws2 = wb.create_sheet(title=sheet_name)
            for ri, row in enumerate(sheet_data, 1):
                for ci, val in enumerate(row, 1):
                    ws2.cell(row=ri, column=ci, value=val)
            # Style header of extra sheet
            hfont = Font(bold=True, color="FFFFFF", size=11)
            hfill = PatternFill(start_color="002060", end_color="002060",
                                fill_type="solid")
            thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))
            for ci in range(1, len(sheet_data[0]) + 1):
                c = ws2.cell(row=1, column=ci)
                c.font = hfont
                c.fill = hfill
                c.alignment = Alignment(horizontal="center", wrap_text=True)
                c.border = thin
            for ri in range(2, len(sheet_data) + 1):
                for ci in range(1, len(sheet_data[0]) + 1):
                    c = ws2.cell(row=ri, column=ci)
                    c.border = thin
                    c.alignment = Alignment(horizontal="center")
            for ci in range(1, len(sheet_data[0]) + 1):
                ws2.column_dimensions[get_column_letter(ci)].width = 22
            ws2.freeze_panes = "A2"

    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  Created: {filename}  ({len(rows)} locations)")


# ===================================================================
# SOV 1: Pacific Retail Group - 47 retail locations in SE US
# ===================================================================
def gen_pacific_retail():
    cities_by_state = {
        "GA": ["Atlanta", "Savannah", "Augusta", "Macon", "Athens",
               "Columbus", "Marietta", "Roswell"],
        "FL": ["Jacksonville", "Tampa", "Orlando", "Miami", "Tallahassee",
               "Fort Myers", "Gainesville", "Pensacola", "Sarasota"],
        "SC": ["Charleston", "Columbia", "Greenville", "Myrtle Beach",
               "Rock Hill"],
        "NC": ["Charlotte", "Raleigh", "Durham", "Greensboro",
               "Wilmington", "Asheville"],
        "TN": ["Nashville", "Memphis", "Knoxville", "Chattanooga",
               "Clarksville"],
        "AL": ["Birmingham", "Huntsville", "Mobile", "Montgomery",
               "Tuscaloosa"],
    }
    bldg_types = [
        ("Shopping Centre", "Retail - Shopping Centre"),
        ("Strip Mall", "Retail - Strip Mall"),
        ("Standalone Retail", "Retail - Standalone"),
        ("Power Centre", "Retail - Power Centre"),
        ("Outlet Store", "Retail - Outlet"),
    ]
    construction = ["Fire Resistive", "Non-Combustible", "Masonry Non-Combustible",
                    "Joisted Masonry"]
    roof_types = ["Built-Up", "Single Ply Membrane", "Metal", "Modified Bitumen"]
    tivs = _distribute_tiv(120_000_000, 47, 0.3, 2.0)
    rows = []
    loc = 1
    states = list(cities_by_state.keys())
    for i in range(47):
        st = states[i % len(states)]
        city = random.choice(cities_by_state[st])
        bt = random.choice(bldg_types)
        ct = random.choice(construction)
        yr = random.randint(1985, 2018)
        stories = random.randint(1, 3)
        sqft = random.randint(8000, 95000)
        sprink = "Y" if random.random() < 0.85 else "N"
        sp_type = random.choice(["Wet", "Dry", "Wet/Dry"]) if sprink == "Y" else "N/A"
        alarm = random.choice(["Central", "Local", "Central"])
        sec = random.choice(["Y", "Y", "N"])
        roof = random.choice(roof_types)
        roof_age = random.randint(2, 25)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        flood = random.choice(["X", "X", "B", "C", "A"])
        wind = random.choice([2, 3, 3, 4])
        eq = random.choice([1, 1, 1, 2])
        coast = round(random.uniform(5, 200), 1)
        fire_dist = round(random.uniform(0.5, 8.0), 1)
        rows.append([
            loc, f"Pacific Retail - {city} {bt[0]}", _rand_street(),
            city, st, _zip_for_state(st), bt[0], bt[1], ct, yr,
            stories, sqft, sprink, sp_type, alarm, sec, roof, roof_age,
            bld, cnt, bi, tiv, flood, wind, eq, coast, fire_dist,
        ])
        loc += 1

    _write_workbook("pacific_retail_sov.xlsx", STANDARD_HEADERS, rows,
                    CURRENCY_COLS_STD)


# ===================================================================
# SOV 2: Acme Industrial - 3 facilities, MISSING columns
# ===================================================================
def gen_acme_industrial():
    facilities = [
        ("Acme Metalworks Plant", "Cleveland", "OH", "44101",
         "Heavy manufacturing - metalworking", "Manufacturing - Metalworking",
         "Joisted Masonry", 1978, 2, 185000),
        ("Acme Assembly Facility", "Akron", "OH", "44304",
         "Assembly plant - automotive parts", "Manufacturing - Assembly",
         "Non-Combustible", 1995, 1, 120000),
        ("Acme Distribution Warehouse", "Pittsburgh", "PA", "15201",
         "Warehouse & distribution", "Warehouse - General",
         "Masonry Non-Combustible", 2002, 1, 95000),
    ]
    tivs = _distribute_tiv(45_000_000, 3, 0.6, 1.4)
    rows = []
    for i, (name, city, st, zp, desc, occ, ct, yr, stories, sqft) in enumerate(facilities):
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        alarm = random.choice(["Central", "Local"])
        sec = "Y"
        roof = random.choice(["Built-Up", "Metal"])
        roof_age = random.randint(5, 30)
        wind = random.choice([3, 4])
        eq = 2
        coast = round(random.uniform(100, 400), 1)
        fire_dist = round(random.uniform(1.0, 5.0), 1)
        rows.append([
            i + 1, name, _rand_street(), city, st, zp,
            desc, occ, ct, yr, stories, sqft,
            # sprinkler columns MISSING
            alarm, sec, roof, roof_age,
            bld, cnt, bi, tiv,
            # flood zone MISSING
            wind, eq, coast, fire_dist,
        ])

    _write_workbook("acme_industrial_sov.xlsx", ACME_HEADERS, rows,
                    CURRENCY_COLS_ACME)


# ===================================================================
# SOV 3: Northern Manufacturing - 1 facility
# ===================================================================
def gen_northern_manufacturing():
    tiv = 28_000_000
    bld, cnt, bi = _split_tiv(tiv)
    rows = [[
        1, "Northern Advanced Manufacturing Center", "4500 Innovation Dr",
        "Allentown", "PA", "18101",
        "Light manufacturing - precision components",
        "Manufacturing - Light", "Fire Resistive", 2015,
        2, 145000, "Y", "Wet", "Central", "Y",
        "Single Ply Membrane", 9, bld, cnt, bi, tiv,
        "X", 4, 2, 185.0, 1.2,
    ]]
    _write_workbook("northern_manufacturing_sov.xlsx", STANDARD_HEADERS,
                    rows, CURRENCY_COLS_STD)


# ===================================================================
# SOV 4: Heritage Hotels - 12 hotels, second sheet for CAT exposure
# ===================================================================
def gen_heritage_hotels():
    fl_locations = [
        ("Heritage Grand Miami Beach", "Miami Beach", "FL", "33139"),
        ("Heritage Resort Key West", "Key West", "FL", "33040"),
        ("Heritage Oceanview Fort Lauderdale", "Fort Lauderdale", "FL", "33301"),
        ("Heritage Bay Naples", "Naples", "FL", "34102"),
        ("Heritage Palm Beach Club", "Palm Beach", "FL", "33480"),
        ("Heritage Emerald Destin", "Destin", "FL", "32541"),
        ("Heritage Sands Clearwater", "Clearwater", "FL", "33755"),
        ("Heritage Historic St Augustine", "St Augustine", "FL", "32084"),
    ]
    carib_locations = [
        ("Heritage Island Bahamas", "Nassau", "Bahamas", "BS-NP"),
        ("Heritage Turks & Caicos Resort", "Providenciales", "Turks & Caicos", "TKCA 1ZZ"),
        ("Heritage St Kitts Plantation Inn", "Basseterre", "St Kitts", "KN-01"),
        ("Heritage Barbados Beach Club", "Bridgetown", "Barbados", "BB15028"),
    ]
    tivs = _distribute_tiv(200_000_000, 12, 0.5, 1.5)
    rows = []
    all_locs = fl_locations + carib_locations
    for i, (name, city, state_country, zp) in enumerate(all_locs):
        yr = random.randint(1995, 2010)
        stories = random.randint(4, 15)
        sqft = random.randint(80000, 350000)
        roof = random.choice(["Concrete Tile", "Clay Tile", "Single Ply Membrane"])
        roof_age = random.randint(3, 18)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        is_coastal_fl = state_country == "FL"
        is_caribbean = state_country not in ("FL",)
        flood = "A" if (is_coastal_fl or is_caribbean) else "B"
        wind = 1 if is_caribbean else (1 if city in ("Key West", "Miami Beach") else 2)
        eq = 1
        coast = round(random.uniform(0.1, 2.5), 1) if (is_coastal_fl or is_caribbean) else round(random.uniform(0.1, 1.0), 1)
        fire_dist = round(random.uniform(1.0, 6.0), 1)
        rows.append([
            i + 1, name, _rand_street(), city, state_country, zp,
            "Luxury hotel & resort", "Hospitality - Luxury Hotel",
            "Fire Resistive", yr, stories, sqft,
            "Y", "Wet", "Central", "Y", roof, roof_age,
            bld, cnt, bi, tiv, flood, wind, eq, coast, fire_dist,
        ])

    # CAT Exposure Summary sheet
    cat_headers = [
        "Location", "City", "State/Country", "TIV",
        "Wind PML (250-yr)", "Wind PML %", "Flood PML (100-yr)",
        "Flood PML %", "Storm Surge Exposure", "Wind Tier",
        "Flood Zone", "Hurricane Deductible %",
    ]
    cat_rows = [cat_headers]
    for r in rows:
        tiv = r[21]
        wind_pml = round(tiv * random.uniform(0.20, 0.45), -3)
        flood_pml = round(tiv * random.uniform(0.08, 0.25), -3)
        cat_rows.append([
            r[1], r[3], r[4], tiv,
            wind_pml, f"{wind_pml/tiv*100:.1f}%",
            flood_pml, f"{flood_pml/tiv*100:.1f}%",
            random.choice(["High", "Very High", "Moderate"]),
            r[23],  # wind tier
            r[22],  # flood zone
            f"{random.choice([2, 3, 5])}%",
        ])
    # Totals in cat
    total_tiv = sum(cr[3] for cr in cat_rows[1:])
    total_wind = sum(cr[4] for cr in cat_rows[1:])
    total_flood = sum(cr[6] for cr in cat_rows[1:])
    cat_rows.append([
        "TOTAL", "", "", total_tiv,
        total_wind, f"{total_wind/total_tiv*100:.1f}%",
        total_flood, f"{total_flood/total_tiv*100:.1f}%",
        "", "", "", "",
    ])

    _write_workbook("heritage_hotels_sov.xlsx", STANDARD_HEADERS, rows,
                    CURRENCY_COLS_STD,
                    extra_sheets={"CAT Exposure Summary": cat_rows})


# ===================================================================
# SOV 5: Western Distribution - 8 warehouses
# ===================================================================
def gen_western_distribution():
    locations = [
        ("Western DC - Denver", "Denver", "CO"),
        ("Western DC - Colorado Springs", "Colorado Springs", "CO"),
        ("Western DC - Salt Lake City", "Salt Lake City", "UT"),
        ("Western DC - Provo", "Provo", "UT"),
        ("Western DC - Phoenix", "Phoenix", "AZ"),
        ("Western DC - Tucson", "Tucson", "AZ"),
        ("Western DC - Las Vegas", "Las Vegas", "NV"),
        ("Western DC - Reno", "Reno", "NV"),
    ]
    tivs = _distribute_tiv(65_000_000, 8, 0.5, 1.5)
    rows = []
    for i, (name, city, st) in enumerate(locations):
        yr = random.randint(2008, 2015)
        stories = 1
        sqft = random.randint(60000, 250000)
        roof = random.choice(["Metal", "Single Ply Membrane", "Built-Up"])
        roof_age = random.randint(3, 15)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        flood = random.choice(["X", "X", "X", "C"])
        wind = 4
        eq = 3 if st in ("UT", "NV") else 2
        coast = round(random.uniform(300, 900), 1)
        fire_dist = round(random.uniform(1.0, 7.0), 1)
        rows.append([
            i + 1, name, _rand_street(), city, st, _zip_for_state(st),
            "Warehouse & distribution center", "Warehouse - Distribution",
            "Non-Combustible", yr, stories, sqft,
            "Y", random.choice(["Wet", "ESFR"]), "Central", "Y",
            roof, roof_age, bld, cnt, bi, tiv, flood, wind, eq, coast,
            fire_dist,
        ])

    _write_workbook("western_distribution_sov.xlsx", STANDARD_HEADERS,
                    rows, CURRENCY_COLS_STD)


# ===================================================================
# SOV 6: Cascade Tech - 2 data centre campuses
# ===================================================================
def gen_cascade_tech():
    campuses = [
        ("Cascade Data Centre - East Campus", "Redmond", "WA", 2019,
         4, 180000),
        ("Cascade Data Centre - West Campus", "Tukwila", "WA", 2020,
         3, 145000),
    ]
    tivs = _distribute_tiv(180_000_000, 2, 0.85, 1.15)
    rows = []
    for i, (name, city, st, yr, stories, sqft) in enumerate(campuses):
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        rows.append([
            i + 1, name, _rand_street(), city, st, _zip_for_state(st),
            "Data centre campus - mission critical IT infrastructure",
            "Technology - Data Centre", "Fire Resistive", yr,
            stories, sqft, "Y", "Pre-Action / Clean Agent",
            "Central", "Y", "Single Ply Membrane", random.randint(2, 5),
            bld, cnt, bi, tiv, "X", 4,
            3,  # earthquake zone - PNW
            round(random.uniform(10, 30), 1),
            round(random.uniform(1.0, 4.0), 1),
        ])

    _write_workbook("cascade_tech_sov.xlsx", STANDARD_HEADERS, rows,
                    CURRENCY_COLS_STD)


# ===================================================================
# SOV 7: Gulf Coast Energy - 5 energy storage facilities
# ===================================================================
def gen_gulf_coast_energy():
    locations = [
        ("Gulf Energy Storage - Houston", "Houston", "TX"),
        ("Gulf Energy Storage - Galveston", "Galveston", "TX"),
        ("Gulf Energy Storage - Corpus Christi", "Corpus Christi", "TX"),
        ("Gulf Energy Storage - Beaumont", "Beaumont", "TX"),
        ("Gulf Energy Storage - Port Arthur", "Port Arthur", "TX"),
    ]
    tivs = _distribute_tiv(95_000_000, 5, 0.6, 1.4)
    rows = []
    for i, (name, city, st) in enumerate(locations):
        yr = random.randint(2016, 2020)
        stories = 1
        sqft = random.randint(35000, 120000)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        flood = random.choice(["A", "A", "B"])
        wind = random.choice([1, 1, 2])
        coast_dist = round(random.uniform(0.5, 25.0), 1)
        rows.append([
            i + 1, name, _rand_street(), city, st, _zip_for_state(st),
            "Energy storage facility - battery / LNG",
            "Energy - Storage Facility", "Non-Combustible", yr,
            stories, sqft, "Y", "Dry / Deluge",
            "Central", "Y", "Metal", random.randint(2, 8),
            bld, cnt, bi, tiv, flood, wind, 1, coast_dist,
            round(random.uniform(2.0, 10.0), 1),
        ])

    _write_workbook("gulf_coast_energy_sov.xlsx", STANDARD_HEADERS, rows,
                    CURRENCY_COLS_STD)


# ===================================================================
# SOV 8: Midwest Ag - 22 grain facilities, DIFFERENT column headers
# ===================================================================
def gen_midwest_ag():
    cities = {
        "MN": ["Minneapolis", "St Paul", "Duluth", "Rochester", "Mankato",
               "Moorhead"],
        "IA": ["Des Moines", "Cedar Rapids", "Davenport", "Sioux City",
               "Waterloo", "Iowa City"],
        "WI": ["Milwaukee", "Madison", "Green Bay", "Appleton", "Oshkosh"],
        "ND": ["Fargo", "Bismarck", "Grand Forks", "Minot", "Williston"],
    }
    descs = [
        ("Grain elevator & storage", "Agriculture - Grain Storage"),
        ("Feed processing plant", "Agriculture - Feed Processing"),
        ("Grain drying facility", "Agriculture - Grain Drying"),
        ("Seed cleaning & storage", "Agriculture - Seed Processing"),
        ("Fertilizer storage", "Agriculture - Chemical Storage"),
    ]
    tivs = _distribute_tiv(35_000_000, 22, 0.4, 1.6)
    # Ensure individual values stay in $800K - $3M range
    for i in range(len(tivs)):
        tivs[i] = max(800_000, min(3_000_000, tivs[i]))
    # Re-scale to hit $35M
    current = sum(tivs)
    factor = 35_000_000 / current
    tivs = [round(t * factor, -3) for t in tivs]

    rows = []
    states = list(cities.keys())
    for i in range(22):
        st = states[i % len(states)]
        city = random.choice(cities[st])
        desc = random.choice(descs)
        yr = random.randint(1985, 2005)
        stories = random.randint(1, 3)
        sqft = random.randint(5000, 40000)
        # Mostly unsprinklered
        sprink = "No" if random.random() < 0.75 else "Yes"
        prot_type = "N/A" if sprink == "No" else random.choice(["Wet", "Dry"])
        alarm = random.choice(["Local", "None", "Local"])
        sec = random.choice(["N", "N", "Y"])
        roof = random.choice(["Metal", "Built-Up", "Metal"])
        roof_age = random.randint(5, 35)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        flood = random.choice(["B", "C", "X"])
        wind = 4
        eq = 1
        coast = round(random.uniform(500, 1200), 1)
        fire_dist = round(random.uniform(3.0, 15.0), 1)
        rows.append([
            i + 1, f"Midwest Ag Facility #{i+1} - {city}",
            _rand_street(), city, st, _zip_for_state(st),
            desc[0], desc[1], "Heavy Timber", yr,
            stories, sqft, sprink, prot_type, alarm, sec,
            roof, roof_age, bld, cnt, bi, tiv,
            flood, wind, eq, coast, fire_dist,
        ])

    _write_workbook("midwest_ag_sov.xlsx", MIDWEST_HEADERS, rows,
                    CURRENCY_COLS_MIDWEST)


# ===================================================================
# SOV 9: Atlantic Seafood - 4 processing plants
# ===================================================================
def gen_atlantic_seafood():
    plants = [
        ("Atlantic Seafood Processing - Gloucester", "Gloucester", "MA",
         "01930", 1988),
        ("Atlantic Seafood Cold Storage - New Bedford", "New Bedford", "MA",
         "02740", 1982),
        ("Atlantic Seafood Packaging - Plymouth", "Plymouth", "MA",
         "02360", 1995),
        ("Atlantic Seafood Distribution - Provincetown", "Provincetown", "MA",
         "02657", 2000),
    ]
    tivs = _distribute_tiv(50_000_000, 4, 0.6, 1.4)
    rows = []
    for i, (name, city, st, zp, yr) in enumerate(plants):
        stories = random.randint(1, 3)
        sqft = random.randint(25000, 85000)
        roof = random.choice(["Built-Up", "Metal", "Modified Bitumen"])
        roof_age = random.randint(5, 25)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        flood = random.choice(["A", "A", "B"])
        wind = random.choice([2, 3])
        coast = round(random.uniform(0.2, 3.0), 1)
        rows.append([
            i + 1, name, _rand_street(), city, st, zp,
            "Seafood processing & cold storage facility",
            "Food Processing - Seafood", "Joisted Masonry", yr,
            stories, sqft, "Y", "Wet", "Central", "Y",
            roof, roof_age, bld, cnt, bi, tiv,
            flood, wind, 2,  # EQ zone - New England
            coast, round(random.uniform(1.5, 5.0), 1),
        ])

    _write_workbook("atlantic_seafood_sov.xlsx", STANDARD_HEADERS, rows,
                    CURRENCY_COLS_STD)


# ===================================================================
# SOV 10: Summit Office - 6 Class A office buildings NY metro
# ===================================================================
def gen_summit_office():
    offices = [
        ("Summit Tower I", "New York", "NY", "10001", 2012, 25, 450000),
        ("Summit Tower II", "New York", "NY", "10036", 2015, 30, 520000),
        ("Summit Brooklyn Heights", "Brooklyn", "NY", "11201", 2016, 18, 280000),
        ("Summit Waterfront", "Jersey City", "NJ", "07302", 2014, 22, 340000),
        ("Summit Stamford Centre", "Stamford", "CT", "06901", 2010, 12, 180000),
        ("Summit White Plains", "White Plains", "NY", "10601", 2018, 10, 155000),
    ]
    tivs = _distribute_tiv(75_000_000, 6, 0.5, 1.5)
    rows = []
    for i, (name, city, st, zp, yr, stories, sqft) in enumerate(offices):
        roof = "Single Ply Membrane"
        roof_age = random.randint(2, 12)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        flood = random.choice(["X", "X", "B"])
        wind = random.choice([3, 4])
        coast_val = round(random.uniform(1.0, 30.0), 1)
        eq = 2
        rows.append([
            i + 1, name, _rand_street(), city, st, zp,
            "Class A commercial office building",
            "Office - Class A", "Fire Resistive", yr,
            stories, sqft, "Y", "Wet", "Central", "Y",
            roof, roof_age, bld, cnt, bi, tiv,
            flood, wind, eq, coast_val,
            round(random.uniform(0.5, 3.0), 1),
        ])

    _write_workbook("summit_office_sov.xlsx", STANDARD_HEADERS, rows,
                    CURRENCY_COLS_STD)


# ===================================================================
# SOV 11: Beacon Light Industrial - 5 buildings NJ (extra)
# ===================================================================
def gen_beacon_light():
    facilities = [
        ("Beacon Light Industrial #1", "Edison", "NJ", "08817", 2005, 1, 55000),
        ("Beacon Light Industrial #2", "Piscataway", "NJ", "08854", 2008, 1, 72000),
        ("Beacon Light Industrial #3", "South Brunswick", "NJ", "08852", 2001, 2, 48000),
        ("Beacon Light Industrial #4", "East Brunswick", "NJ", "08816", 2010, 1, 63000),
        ("Beacon Light Industrial #5", "North Brunswick", "NJ", "08902", 2003, 1, 58000),
    ]
    tivs = _distribute_tiv(40_000_000, 5, 0.6, 1.4)
    rows = []
    for i, (name, city, st, zp, yr, stories, sqft) in enumerate(facilities):
        ct = random.choice(["Non-Combustible", "Masonry Non-Combustible"])
        sprink = "Y"
        sp_type = random.choice(["Wet", "Dry"])
        roof = random.choice(["Metal", "Single Ply Membrane"])
        roof_age = random.randint(3, 18)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        flood = random.choice(["X", "B", "C"])
        wind = random.choice([3, 4])
        eq = 2
        coast = round(random.uniform(15, 50), 1)
        rows.append([
            i + 1, name, _rand_street(), city, st, zp,
            "Light industrial / flex space",
            "Industrial - Light Manufacturing", ct, yr,
            stories, sqft, sprink, sp_type, "Central", "Y",
            roof, roof_age, bld, cnt, bi, tiv,
            flood, wind, eq, coast,
            round(random.uniform(1.0, 5.0), 1),
        ])

    _write_workbook("beacon_light_sov.xlsx", STANDARD_HEADERS, rows,
                    CURRENCY_COLS_STD)


# ===================================================================
# SOV 12: Golden Gate Retail - 15 locations CA (extra)
# ===================================================================
def gen_golden_gate_retail():
    ca_cities = [
        "San Francisco", "Los Angeles", "San Diego", "Sacramento",
        "San Jose", "Oakland", "Fresno", "Long Beach", "Santa Ana",
        "Anaheim", "Irvine", "Pasadena", "Berkeley", "Palo Alto",
        "Santa Monica",
    ]
    bldg_types = [
        ("Shopping Centre", "Retail - Shopping Centre"),
        ("Strip Mall", "Retail - Strip Mall"),
        ("Standalone Retail", "Retail - Standalone"),
        ("Mixed-Use Retail/Office", "Retail - Mixed Use"),
    ]
    tivs = _distribute_tiv(85_000_000, 15, 0.4, 1.6)
    rows = []
    for i in range(15):
        city = ca_cities[i]
        bt = random.choice(bldg_types)
        ct = random.choice(["Fire Resistive", "Non-Combustible",
                            "Masonry Non-Combustible"])
        yr = random.randint(1990, 2020)
        stories = random.randint(1, 4)
        sqft = random.randint(10000, 80000)
        sprink = "Y" if random.random() < 0.90 else "N"
        sp_type = random.choice(["Wet", "Dry"]) if sprink == "Y" else "N/A"
        alarm = random.choice(["Central", "Local"])
        sec = random.choice(["Y", "Y", "N"])
        roof = random.choice(["Built-Up", "Single Ply Membrane", "Clay Tile"])
        roof_age = random.randint(2, 20)
        tiv = tivs[i]
        bld, cnt, bi = _split_tiv(tiv)
        flood = random.choice(["X", "X", "B", "C"])
        wind = 4
        eq = random.choice([4, 4, 5])  # CA high earthquake zone
        coast = round(random.uniform(1, 80), 1)
        rows.append([
            i + 1, f"Golden Gate Retail - {city}", _rand_street(),
            city, "CA", _zip_for_state("CA"),
            bt[0], bt[1], ct, yr, stories, sqft,
            sprink, sp_type, alarm, sec, roof, roof_age,
            bld, cnt, bi, tiv, flood, wind, eq, coast,
            round(random.uniform(0.5, 6.0), 1),
        ])

    _write_workbook("golden_gate_retail_sov.xlsx", STANDARD_HEADERS, rows,
                    CURRENCY_COLS_STD)


# ===================================================================
# Main
# ===================================================================
if __name__ == "__main__":
    print("Generating SOV schedules...")
    gen_pacific_retail()
    gen_acme_industrial()
    gen_northern_manufacturing()
    gen_heritage_hotels()
    gen_western_distribution()
    gen_cascade_tech()
    gen_gulf_coast_energy()
    gen_midwest_ag()
    gen_atlantic_seafood()
    gen_summit_office()
    gen_beacon_light()
    gen_golden_gate_retail()
    print(f"\nDone. 12 SOV files written to:\n  {OUTPUT_DIR}")
